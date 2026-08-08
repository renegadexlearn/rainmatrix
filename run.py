#!/usr/bin/env python3
"""
Mindoro Rain Matrix — Flask Web App (with SQLite cache)

- Loads places from places.json
- Visit:
    http://127.0.0.1:5000/                 -> shows today's matrix (Asia/Manila)
    http://127.0.0.1:5000/?date=2025-12-31 -> shows that day's matrix

Query params:
    date=YYYY-MM-DD
    tz=Asia/Manila
    country=PH
    model=ecmwf_ifs
    nocache=1   -> bypass SQLite HTML cache (useful while tweaking CSS/colors)

Day range policy:
    Allowed dates are only from "today" (in tz) up to 4 days in the future.
"""

from __future__ import annotations

import os
import sqlite3
import json
import calendar as calendar_lib
import html as html_lib
import statistics
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, date, timedelta, timezone
from typing import Dict, List, Optional, Tuple
from urllib.parse import urlencode

import requests
from flask import Flask, request, Response
from openpyxl import load_workbook

try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None  # type: ignore

# ---------------- CONFIG ----------------

DEFAULT_TZ = "Asia/Manila"
DEFAULT_COUNTRY = "PH"
DEFAULT_MODEL = "ecmwf_ifs"
DEFAULT_PLACES_FILE = "places.json"

DAY_START_HOUR = 6
DAY_END_HOUR = 18

CLEAR_MAX = 25.0
PARTLY_MAX = 60.0

LIGHT_MAX = 2.5
MODERATE_MAX = 7.5

SCALE_MAX_MM = 7.0
SOLAR_SCALE_MAX_WM2 = 1000.0
SOLAR_TOTAL_CACHE_VERSION = "solar-total-v2"
BATTERY_NOMINAL_VOLTAGE = 51.2
RAIN_RETAINED_ENERGY_FLOOR = 0.12
SMART_RAIN_MIN_MM = 0.05
SMART_RAIN_MIN_PROBABILITY = 30.0
MAX_INVERTER_SAMPLE_GAP_MINUTES = 20.0
INVERTER_DAY_COMPLETE_HOUR = 18
SOLAR_CALIBRATION_MIN_DAYS = 3
SOLAR_CALIBRATION_LOOKBACK_DAYS = 90

SKYBLUE = "#87CEEB"
VIOLET = "#8A2BE2"

# Time column palette (from your image)
TC_1 = "#FFF2BD"
TC_2 = "#F4D797"
TC_3 = "#EBB58A"
TC_4 = "#DA7F7D"
TC_5 = "#B5728E"
TC_6 = "#776E99"

# Day changer range: today .. today+4
FUTURE_DAYS_ALLOWED = 4

WEATHER_ICONS_NO_RAIN = {
    "clear_day": "☀️",
    "clear_night": "🌙",
    "partly_day": "🌤️",
    "partly_night": "🌙☁️",
    "cloudy": "☁️",
}

# SQLite cache DB path (override via env var)
CACHE_DB_PATH = os.path.abspath(os.environ.get("RAIN_CACHE_DB", "rain_cache.sqlite3"))

# Cache policy
CACHE_TTL_SECONDS = 60 * 60          # 1 hour
CACHE_RETENTION_DAYS = 2             # delete rows older than 2 days

# ---------------- DATA ----------------

@dataclass
class Place:
    label: str
    query: str
    lat: float
    lon: float
    admin: str = ""

# ---------------- HELPERS ----------------

def _to_float(x, default: float = 0.0) -> float:
    try:
        return float(x)
    except (TypeError, ValueError):
        return default


def is_day(dt: datetime) -> bool:
    return DAY_START_HOUR <= dt.hour < DAY_END_HOUR


def precip_display(p_mm: float) -> str:
    return "-" if _to_float(p_mm) <= 0 else f"{p_mm:.1f}"


def weather_icon(cloudcover_pct: float, precipitation_mm: float, dt: datetime) -> str:
    cc = _to_float(cloudcover_pct)
    p = _to_float(precipitation_mm)
    dayflag = is_day(dt)

    if p > 0:
        if p <= LIGHT_MAX:
            return "🌦️" if dayflag else "🌧️"
        elif p <= MODERATE_MAX:
            return "🌧️"
        return "⛈️"

    if cc <= CLEAR_MAX:
        return WEATHER_ICONS_NO_RAIN["clear_day"] if dayflag else WEATHER_ICONS_NO_RAIN["clear_night"]
    elif cc <= PARTLY_MAX:
        return WEATHER_ICONS_NO_RAIN["partly_day"] if dayflag else WEATHER_ICONS_NO_RAIN["partly_night"]
    return WEATHER_ICONS_NO_RAIN["cloudy"]


def _hex_to_rgb(h: str) -> Tuple[int, int, int]:
    h = h.lstrip("#")
    return int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)


def _rgb_to_hex(r: int, g: int, b: int) -> str:
    return f"#{r:02X}{g:02X}{b:02X}"


def _clamp01(x: float) -> float:
    return 0.0 if x < 0.0 else 1.0 if x > 1.0 else x


def _smoothstep(t: float) -> float:
    t = _clamp01(t)
    return t * t * (3.0 - 2.0 * t)


def _lerp_color(c0: str, c1: str, t: float) -> str:
    t = _smoothstep(t)
    r0, g0, b0 = _hex_to_rgb(c0)
    r1, g1, b1 = _hex_to_rgb(c1)
    return _rgb_to_hex(
        round(r0 + (r1 - r0) * t),
        round(g0 + (g1 - g0) * t),
        round(b0 + (b1 - b0) * t),
    )


def precip_bg_color(p_mm: float) -> str:
    p = max(0.0, _to_float(p_mm))
    if p >= SCALE_MAX_MM:
        return VIOLET

    t = p / SCALE_MAX_MM
    r0, g0, b0 = _hex_to_rgb(SKYBLUE)
    r1, g1, b1 = _hex_to_rgb(VIOLET)

    return _rgb_to_hex(
        round(r0 + (r1 - r0) * t),
        round(g0 + (g1 - g0) * t),
        round(b0 + (b1 - b0) * t),
    )


def solar_bg_color(watts_m2: float) -> str:
    w = max(0.0, _to_float(watts_m2))
    t = min(w / SOLAR_SCALE_MAX_WM2, 1.0)

    if t <= 0.5:
        return _lerp_color("#F1F5F9", "#FDE68A", t / 0.5)
    return _lerp_color("#FDE68A", "#F59E0B", (t - 0.5) / 0.5)


def solar_display(watts_m2: float) -> str:
    w = _to_float(watts_m2)
    return "-" if w <= 0 else f"{round(w):.0f}"


def solar_total_display(watts_m2: float) -> str:
    return f"{round(max(0.0, _to_float(watts_m2))):.0f}"


def time_bg_color(dt: datetime) -> str:
    """
    Smooth day/night gradient for Time column using your palette.
    """
    h = dt.hour + (dt.minute / 60.0)

    stops: List[Tuple[float, str]] = [
        (0.0,  TC_6),
        (4.5,  TC_5),
        (6.5,  TC_4),
        (8.5,  TC_3),
        (11.0, TC_2),
        (13.0, TC_1),
        (15.5, TC_2),
        (18.0, TC_3),
        (19.5, TC_4),
        (21.5, TC_5),
        (24.0, TC_6),
    ]

    for (h0, c0), (h1, c1) in zip(stops, stops[1:]):
        if h0 <= h <= h1:
            if h1 == h0:
                return c1
            t = (h - h0) / (h1 - h0)
            return _lerp_color(c0, c1, t)

    return TC_6


def hour_label(dt: datetime) -> str:
    return dt.strftime("%H:00")


def place_label_from_query(q: str) -> str:
    return q.split(",")[0].strip()


def read_places_file(path: str, list_id: Optional[str] = None) -> Tuple[List[Place], List[Dict[str, str]], str]:
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    default_list_id = data.get("default_list", "")
    lists_data = data.get("lists", [])

    if not lists_data:
        raise ValueError(f"No lists found in {path}.")

    if list_id is None:
        list_id = default_list_id

    selected_list = next((lst for lst in lists_data if lst["id"] == list_id), None)
    if selected_list is None:
        selected_list = lists_data[0]
        list_id = selected_list["id"]

    places: List[Place] = []
    for p in selected_list.get("places", []):
        places.append(
            Place(
                label=p["label"],
                query=p["label"],
                lat=float(p["lat"]),
                lon=float(p["lon"]),
            )
        )

    lists_info = [{"id": lst["id"], "name": lst["name"]} for lst in lists_data]

    return places, lists_info, list_id


def places_signature(path: str, list_id: str) -> str:
    st = os.stat(path)
    return f"{int(st.st_mtime)}:{st.st_size}:{list_id}"


def read_config_file(path: str) -> Dict[str, object]:
    if not os.path.exists(path):
        raise FileNotFoundError(path)

    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    if "lists" not in data:
        data["lists"] = []
    if "sites" not in data:
        data["sites"] = []

    return data


def read_sites_file(path: str) -> List[Dict[str, object]]:
    data = read_config_file(path)
    sites = data.get("sites", [])
    return sites if isinstance(sites, list) else []


_INVERTER_FILE_CACHE: Dict[str, Tuple[Tuple[int, int], Dict[str, object]]] = {}


def _parse_inverter_time(value: object) -> Optional[datetime]:
    if isinstance(value, datetime):
        return value.replace(tzinfo=None)
    if not value:
        return None
    text = str(value).strip()
    for time_format in ("%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, time_format)
        except ValueError:
            continue
    return None


def inverter_actuals_for_file(path: str) -> Dict[str, object]:
    empty = {"hourly": {}, "daily": {}, "last_reading": None, "source": ""}
    if not path:
        return empty

    resolved_path = os.path.abspath(path)
    try:
        stat = os.stat(resolved_path)
    except OSError:
        return empty

    signature = (stat.st_mtime_ns, stat.st_size)
    cached = _INVERTER_FILE_CACHE.get(resolved_path)
    if cached and cached[0] == signature:
        return cached[1]

    points_by_time: Dict[datetime, float] = {}
    workbook = load_workbook(resolved_path, read_only=True, data_only=True)
    try:
        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            headers = next(rows, None)
            if not headers:
                continue
            normalized_headers = [
                str(value or "").strip().lower().replace("（", "(").replace("）", ")")
                for value in headers
            ]
            time_index = next((i for i, value in enumerate(normalized_headers) if value == "time"), None)
            solar_index = next(
                (i for i, value in enumerate(normalized_headers) if value.startswith("solar power")),
                None,
            )
            weather_index = next(
                (i for i, value in enumerate(normalized_headers) if value == "weather"),
                None,
            )
            soc_index = next(
                (i for i, value in enumerate(normalized_headers) if value.startswith("soc")),
                None,
            )
            if time_index is None or solar_index is None:
                continue

            for row in rows:
                reading_time = _parse_inverter_time(row[time_index])
                if reading_time is None:
                    continue
                weather_value = row[weather_index] if weather_index is not None else None
                soc_value = row[soc_index] if soc_index is not None else 1
                if weather_value and soc_value is None:
                    continue
                try:
                    solar_power_kw = max(0.0, float(row[solar_index] or 0.0))
                except (TypeError, ValueError):
                    continue
                points_by_time[reading_time] = solar_power_kw
    finally:
        workbook.close()

    points = sorted(points_by_time.items())
    hourly_energy: Dict[datetime, float] = defaultdict(float)
    hourly_coverage_seconds: Dict[datetime, float] = defaultdict(float)
    for (start, start_kw), (end, end_kw) in zip(points, points[1:]):
        interval_seconds = (end - start).total_seconds()
        if interval_seconds <= 0 or interval_seconds > MAX_INVERTER_SAMPLE_GAP_MINUTES * 60.0:
            continue
        cursor = start
        while cursor < end:
            hour_start = cursor.replace(minute=0, second=0, microsecond=0)
            boundary = min(end, hour_start + timedelta(hours=1))
            start_fraction = (cursor - start).total_seconds() / interval_seconds
            end_fraction = (boundary - start).total_seconds() / interval_seconds
            cursor_kw = start_kw + ((end_kw - start_kw) * start_fraction)
            boundary_kw = start_kw + ((end_kw - start_kw) * end_fraction)
            segment_seconds = (boundary - cursor).total_seconds()
            hourly_energy[hour_start] += ((cursor_kw + boundary_kw) / 2.0) * (segment_seconds / 3600.0)
            hourly_coverage_seconds[hour_start] += segment_seconds
            cursor = boundary

    hourly = {
        hour: {
            "kwh": max(0.0, energy),
            "coverage_minutes": min(hourly_coverage_seconds[hour] / 60.0, 60.0),
        }
        for hour, energy in hourly_energy.items()
    }
    last_by_day: Dict[date, datetime] = {}
    for reading_time, _power in points:
        last_by_day[reading_time.date()] = max(
            last_by_day.get(reading_time.date(), reading_time),
            reading_time,
        )
    daily = {
        reading_day: {
            "kwh": sum(
                _to_float(values.get("kwh"))
                for hour, values in hourly.items()
                if hour.date() == reading_day
            ),
            "last_reading": last_reading,
            "complete": last_reading.hour >= INVERTER_DAY_COMPLETE_HOUR,
        }
        for reading_day, last_reading in last_by_day.items()
    }
    result = {
        "hourly": hourly,
        "daily": daily,
        "last_reading": points[-1][0] if points else None,
        "source": os.path.basename(resolved_path),
    }
    _INVERTER_FILE_CACHE[resolved_path] = (signature, result)
    return result


def safe_now_date_in_tz(tz_name: str) -> date:
    return safe_now_in_tz(tz_name).date()


def safe_now_in_tz(tz_name: str) -> datetime:
    if ZoneInfo is None:
        return datetime.now()

    try:
        return datetime.now(ZoneInfo(tz_name)).replace(tzinfo=None)
    except Exception:
        return datetime.now()


def build_url(base_params: Dict[str, str], new_date: Optional[date]) -> str:
    params = dict(base_params)
    if new_date is None:
        params.pop("date", None)
    else:
        params["date"] = new_date.isoformat()
    return "/?" + urlencode(params)


def normalized_view(view: Optional[str]) -> str:
    return "solar" if view == "solar" else "rain"


def normalize_efficiency_factor(value: float) -> float:
    factor = _to_float(value, 0.80)
    if factor > 1.0:
        factor = factor / 100.0
    return max(0.0, min(factor, 1.0))


def daily_solar_production_kwh(system_kw: float, irradiation_wh_m2: float, efficiency_factor: float) -> float:
    return max(0.0, system_kw) * (max(0.0, irradiation_wh_m2) / 1000.0) * normalize_efficiency_factor(efficiency_factor)


def rain_retained_energy_factor(precipitation_mm: float, precipitation_probability_pct: float) -> float:
    """Return the conservative share of forecast solar energy retained during rain."""
    rain_probability = min(max(_to_float(precipitation_probability_pct) / 100.0, 0.0), 1.0)
    rain_intensity = min(max(_to_float(precipitation_mm), 0.0) / 2.0, 1.0)

    # Probability handles scattered tropical rain that models often understate in mm.
    # Intensity supplies the extra loss once meaningful rainfall is forecast.
    loss = (0.88 * rain_probability) + (0.08 * rain_intensity)
    return max(RAIN_RETAINED_ENERGY_FLOOR, min(1.0 - loss, 1.0))


def smart_hourly_rain_adjustment(
        precipitation_mm: float,
        precipitation_probability_pct: float,
    ) -> Tuple[str, float]:
    """Classify one forecast hour and return its retained solar-energy share."""
    precipitation = max(0.0, _to_float(precipitation_mm))
    probability = min(max(_to_float(precipitation_probability_pct), 0.0), 100.0)
    if precipitation < SMART_RAIN_MIN_MM or probability < SMART_RAIN_MIN_PROBABILITY:
        return "dry", 1.0

    probability_weight = probability / 100.0
    intensity_weight = min(precipitation / 2.0, 1.0)
    # Inverter readings from CAL Tree of Life show that satellite radiation
    # remains optimistic during sustained tropical rain. Weight probability
    # strongly, then use hourly intensity to deepen the loss.
    loss = (0.80 * probability_weight) + (0.20 * intensity_weight)
    retained = max(RAIN_RETAINED_ENERGY_FLOOR, min(1.0 - loss, 1.0))

    if probability >= 80.0 and precipitation >= 1.0:
        return "heavy", retained
    if probability >= 60.0 and precipitation >= 0.3:
        return "likely", retained
    return "light", retained


def hourly_solar_production_kwh(
        system_kw: float,
        radiation_w_m2: float,
        efficiency_factor: float,
        precipitation_mm: float = 0.0,
        precipitation_probability_pct: float = 0.0,
        rain_effect: bool = False,
        calibration_factor: float = 1.0,
    ) -> float:
    production = daily_solar_production_kwh(system_kw, radiation_w_m2, efficiency_factor)
    if rain_effect:
        production *= rain_retained_energy_factor(precipitation_mm, precipitation_probability_pct)
    return max(0.0, production * max(0.0, _to_float(calibration_factor, 1.0)))


def production_bar_color(kwh: float, max_kwh: float) -> str:
    if max_kwh <= 0:
        return "#D1D5DB"
    t = min(max(kwh, 0.0) / max_kwh, 1.0)
    if t < 0.35:
        return _lerp_color("#D1D5DB", "#FDE68A", t / 0.35)
    if t < 0.70:
        return _lerp_color("#FDE68A", "#FBBF24", (t - 0.35) / 0.35)
    return _lerp_color("#FBBF24", "#F97316", (t - 0.70) / 0.30)


def nice_solar_axis_max(max_kwh: float, step: int = 25) -> int:
    if max_kwh <= 0:
        return step
    return max(step, int(((max_kwh + step - 0.000001) // step) * step))

# ---------------- CLIENT ----------------

class OpenMeteoClient:
    def __init__(self, timeout: int = 20):
        self.timeout = timeout
        self.session = requests.Session()

    def geocode(self, query: str, country_code: Optional[str]) -> Optional[Place]:
        name_only = query.split(",")[0].strip()

        params = {
            "name": name_only,
            "count": 10,
            "language": "en",
            "format": "json",
        }
        if country_code:
            params["country_code"] = country_code.upper()

        r = self.session.get(
            "https://geocoding-api.open-meteo.com/v1/search",
            params=params,
            timeout=self.timeout,
        )
        r.raise_for_status()
        results = r.json().get("results") or []
        if not results:
            return None

        best = max(results, key=lambda x: x.get("population") or 0)

        return Place(
            label=place_label_from_query(query),
            query=query,
            lat=best["latitude"],
            lon=best["longitude"],
        )

    def hourly_forecast(self, lat: float, lon: float, tz: str, model: str):
        params = {
            "latitude": lat,
            "longitude": lon,
            "hourly": "precipitation,precipitation_probability,cloud_cover,shortwave_radiation",
            "timezone": tz,
            "forecast_days": 7,
            "models": model,
        }
        r = self.session.get(
            "https://api.open-meteo.com/v1/forecast",
            params=params,
            timeout=self.timeout,
        )
        r.raise_for_status()
        h = r.json()["hourly"]
        return {
            "time": [datetime.fromisoformat(t) for t in h["time"]],
            "precip": h["precipitation"],
            "pop": h.get("precipitation_probability") or [0] * len(h["time"]),
            "cloud": h["cloud_cover"],
            "solar": h.get("shortwave_radiation") or [0] * len(h["time"]),
        }

    def satellite_radiation_today(self, lat: float, lon: float, tz: str):
        params = {
            "latitude": lat,
            "longitude": lon,
            "hourly": "shortwave_radiation",
            "timezone": tz,
            "forecast_days": 1,
        }
        r = self.session.get(
            "https://satellite-api.open-meteo.com/v1/archive",
            params=params,
            timeout=self.timeout,
        )
        r.raise_for_status()
        h = r.json()["hourly"]
        return {
            "time": [datetime.fromisoformat(t) for t in h["time"]],
            "solar": h.get("shortwave_radiation") or [0] * len(h["time"]),
        }

    def hybrid_solar_forecast(self, lat: float, lon: float, tz: str, model: str):
        hourly = self.hourly_forecast(lat, lon, tz, model)
        now = safe_now_in_tz(tz)
        completed_hour = now.replace(minute=0, second=0, microsecond=0)
        hourly["solar_source"] = ["forecast"] * len(hourly["time"])

        try:
            satellite = self.satellite_radiation_today(lat, lon, tz)
        except Exception:
            return hourly

        satellite_by_time = {
            t: max(0.0, _to_float(solar))
            for t, solar in zip(satellite["time"], satellite["solar"])
            if t.date() == now.date() and t <= completed_hour and solar is not None
        }
        if not satellite_by_time:
            return hourly

        hourly["solar"] = [
            satellite_by_time.get(t, max(0.0, _to_float(solar)))
            if t.date() == now.date()
            else max(0.0, _to_float(solar))
            for t, solar in zip(hourly["time"], hourly["solar"])
        ]
        hourly["solar_source"] = [
            "nowcast" if t in satellite_by_time else "forecast"
            for t in hourly["time"]
        ]
        return hourly

    def add_rain_model_scenarios(
            self,
            hourly: Dict[str, object],
            lat: float,
            lon: float,
            tz: str,
            primary_model: str = DEFAULT_MODEL,
        ):
        if hourly.get("rain_solar"):
            return hourly

        primary_solar = hourly.get("solar") or []
        model_series = [(primary_model, hourly)]
        for backup_model in ("gfs_global", "icon_seamless", "gem_global"):
            try:
                model_series.append(
                    (backup_model, self.hourly_forecast(lat, lon, tz, backup_model))
                )
            except Exception:
                continue

        model_labels = {
            "ecmwf_ifs": "ECMWF",
            "best_match": "Best match",
            "gfs_global": "GFS",
            "icon_seamless": "ICON",
            "gem_global": "GEM",
        }

        def hourly_candidates(model_name: str, series: Dict[str, object]):
            candidates = {}
            solar_sources = series.get("solar_source") or ["forecast"] * len(series["time"])
            for t, solar, precipitation, probability, solar_source in zip(
                series["time"],
                series["solar"],
                series["precip"],
                series["pop"],
                solar_sources,
            ):
                if solar is None:
                    continue
                raw_solar = max(0.0, _to_float(solar))
                rain_mm = max(0.0, _to_float(precipitation))
                rain_pop = min(max(_to_float(probability), 0.0), 100.0)
                rain_level, retained_factor = smart_hourly_rain_adjustment(rain_mm, rain_pop)
                candidates[t] = {
                    "model": model_labels.get(model_name, model_name.replace("_", " ").title()),
                    "raw": raw_solar,
                    "adjusted": raw_solar * retained_factor,
                    "precip": rain_mm,
                    "pop": rain_pop,
                    "rain_level": rain_level,
                    "retained_factor": retained_factor,
                    "source": solar_source,
                }
            return candidates

        candidates_by_model = [
            hourly_candidates(model_name, series)
            for model_name, series in model_series
        ]

        expected_solar = []
        conservative_solar = []
        upper_solar = []
        scenario_used = []
        hourly_precip = []
        hourly_pop = []
        selected_models = []
        interpretations = []
        rain_levels = []
        retained_factors = []
        primary_sources = hourly.get("solar_source") or ["forecast"] * len(hourly["time"])
        for t, primary, primary_source in zip(hourly["time"], primary_solar, primary_sources):
            candidates = [model_values[t] for model_values in candidates_by_model if t in model_values]
            if primary_source == "nowcast" and candidates:
                selected = candidates[0]
                rain_mm = selected["precip"]
                rain_pop = selected["pop"]
                rain_level, retained_factor = smart_hourly_rain_adjustment(rain_mm, rain_pop)
                expected = selected["raw"] * retained_factor
                scenario_values = [expected]
                model_label = "Satellite nowcast"
                interpretation = (
                    "Satellite sunlight is not site meter output; hourly rain penalty applied."
                    if retained_factor < 1.0
                    else "Satellite sunlight used; the hourly rain signal is too small for a penalty."
                )
            elif candidates:
                rain_mm = statistics.median(candidate["precip"] for candidate in candidates)
                rain_pop = statistics.median(candidate["pop"] for candidate in candidates)
                rain_level, consensus_factor = smart_hourly_rain_adjustment(rain_mm, rain_pop)

                if rain_level == "dry":
                    scenario_values = [candidate["raw"] for candidate in candidates]
                    target = statistics.median(scenario_values)
                    selected = min(candidates, key=lambda candidate: (abs(candidate["raw"] - target), candidate["raw"]))
                    expected = selected["raw"]
                    retained_factor = 1.0
                    interpretation = "Rain signal is too small; solar forecast used without a rain penalty."
                else:
                    scenario_values = [candidate["raw"] * consensus_factor for candidate in candidates]
                    ordered = sorted(candidates, key=lambda candidate: candidate["raw"])
                    if rain_level == "heavy":
                        selected = ordered[0]
                        interpretation = "Heavy rain is likely; the lowest adjusted model is used."
                    elif rain_level == "likely":
                        selected = ordered[max(0, (len(ordered) - 1) // 2)]
                        interpretation = "Rain is likely; a lower-middle adjusted model is used."
                    else:
                        target = statistics.median(scenario_values)
                        selected = min(
                            candidates,
                            key=lambda candidate: (
                                abs((candidate["raw"] * consensus_factor) - target),
                                candidate["raw"],
                            ),
                        )
                        interpretation = "Light or uncertain rain; the middle adjusted model is used."
                    expected = selected["raw"] * consensus_factor
                    retained_factor = consensus_factor
                model_label = selected["model"]
                if rain_level != "dry" and retained_factor >= 0.995:
                    interpretation = (
                        "Rain signals disagree; the selected model needs no extra rain penalty."
                    )
            else:
                expected = max(0.0, _to_float(primary))
                scenario_values = [expected]
                rain_mm = 0.0
                rain_pop = 0.0
                rain_level = "dry"
                retained_factor = 1.0
                model_label = model_labels.get(primary_model, "Primary")
                interpretation = "Only the primary solar forecast is available."

            expected_solar.append(expected)
            conservative_solar.append(min(scenario_values))
            upper_solar.append(max(scenario_values))
            scenario_used.append(
                len(scenario_values) > 1 and (max(scenario_values) - min(scenario_values)) > 0.01
            )
            hourly_precip.append(rain_mm)
            hourly_pop.append(rain_pop)
            selected_models.append(model_label)
            interpretations.append(interpretation)
            rain_levels.append(rain_level)
            retained_factors.append(retained_factor)

        hourly["rain_solar"] = expected_solar
        hourly["rain_conservative_solar"] = conservative_solar
        hourly["rain_upper_solar"] = upper_solar
        hourly["rain_consensus_used"] = scenario_used
        hourly["rain_model_count"] = len(candidates_by_model)
        hourly["rain_adjustment_precip"] = hourly_precip
        hourly["rain_adjustment_pop"] = hourly_pop
        hourly["rain_selected_model"] = selected_models
        hourly["rain_interpretation"] = interpretations
        hourly["rain_level"] = rain_levels
        hourly["rain_retained_factor"] = retained_factors
        return hourly


# ---------------- CACHE (SQLite) ----------------

def cache_connect() -> sqlite3.Connection:
    conn = sqlite3.connect(CACHE_DB_PATH, timeout=10)
    conn.execute("PRAGMA journal_mode=WAL;")
    conn.execute("PRAGMA synchronous=NORMAL;")
    return conn


def ensure_solar_history_schema(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS solar_production_history (
          site_id                TEXT NOT NULL,
          production_date        TEXT NOT NULL,
          base_forecast_kwh      REAL,
          rain_forecast_kwh      REAL,
          base_display_kwh       REAL,
          rain_display_kwh       REAL,
          source_type            TEXT,
          last_queried_at         TEXT,
          actual_kwh             REAL,
          forecast_recorded_at   TEXT,
          actual_recorded_at     TEXT,
          PRIMARY KEY (site_id, production_date)
        )
        """
    )
    history_columns = {
        row[1]
        for row in conn.execute("PRAGMA table_info(solar_production_history)").fetchall()
    }
    for column, column_type in (
        ("base_display_kwh", "REAL"),
        ("rain_display_kwh", "REAL"),
        ("source_type", "TEXT"),
        ("last_queried_at", "TEXT"),
    ):
        if column in history_columns:
            continue
        try:
            conn.execute(f"ALTER TABLE solar_production_history ADD COLUMN {column} {column_type}")
        except sqlite3.OperationalError as error:
            if "duplicate column name" not in str(error).lower():
                raise


def cache_init() -> None:
    with cache_connect() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS rain_cache (
              query_date  TEXT NOT NULL,
              target_date TEXT NOT NULL,
              tz          TEXT NOT NULL,
              country     TEXT NOT NULL,
              model       TEXT NOT NULL,
              places_sig  TEXT NOT NULL,
              html        TEXT NOT NULL,
              created_at  TEXT NOT NULL,
              PRIMARY KEY (query_date, target_date, tz, country, model, places_sig)
            )
            """
        )
        ensure_solar_history_schema(conn)
    cache_prune()


def cache_prune() -> None:
    """Deletes cache rows older than CACHE_RETENTION_DAYS (based on created_at, UTC)."""
    with cache_connect() as conn:
        conn.execute(
            """
            DELETE FROM rain_cache
            WHERE created_at < datetime('now', ?)
            """,
            (f"-{CACHE_RETENTION_DAYS} days",),
        )


def cache_get(query_date: str, target_date: str, tz: str, country: str, model: str, places_sig: str) -> Optional[str]:
    with cache_connect() as conn:
        row = conn.execute(
            """
            SELECT html, created_at
            FROM rain_cache
            WHERE query_date=? AND target_date=? AND tz=? AND country=? AND model=? AND places_sig=?
            """,
            (query_date, target_date, tz, country, model, places_sig),
        ).fetchone()

        if not row:
            return None

        html, created_at = row

        # created_at stored as UTC "YYYY-MM-DD HH:MM:SS"
        try:
            created_dt = datetime.strptime(created_at, "%Y-%m-%d %H:%M:%S")
        except Exception:
            # If parsing fails, treat as expired to force refresh
            return None

        age = datetime.now(timezone.utc).replace(tzinfo=None) - created_dt
        if age.total_seconds() > CACHE_TTL_SECONDS:
            return None

        return html


def cache_put(query_date: str, target_date: str, tz: str, country: str, model: str, places_sig: str, html: str) -> None:
    with cache_connect() as conn:
        conn.execute(
            """
            INSERT OR REPLACE INTO rain_cache
              (query_date, target_date, tz, country, model, places_sig, html, created_at)
            VALUES
              (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                query_date,
                target_date,
                tz,
                country,
                model,
                places_sig,
                html,
                datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )


def solar_history_record_forecasts(site_id: str, daily_rows: List[Dict[str, object]]) -> None:
    if not site_id:
        return

    recorded_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    rows = []
    for row in daily_rows:
        production_day = row.get("date")
        if not isinstance(production_day, date):
            continue
        rows.append(
            (
                site_id,
                production_day.isoformat(),
                max(0.0, _to_float(row.get("base_forecast_kwh"))),
                max(0.0, _to_float(row.get("rain_forecast_kwh"))),
                max(0.0, _to_float(row.get("base_display_kwh"))),
                max(0.0, _to_float(row.get("rain_display_kwh"))),
                str(row.get("source_type") or "forecast"),
                (
                    max(0.0, _to_float(row.get("inverter_actual_kwh")))
                    if row.get("inverter_actual_complete")
                    else None
                ),
                recorded_at,
                recorded_at,
                recorded_at if row.get("inverter_actual_complete") else None,
            )
        )

    if not rows:
        return

    with cache_connect() as conn:
        ensure_solar_history_schema(conn)
        conn.executemany(
            """
            INSERT INTO solar_production_history
              (site_id, production_date, base_forecast_kwh, rain_forecast_kwh,
               base_display_kwh, rain_display_kwh, source_type,
               actual_kwh, forecast_recorded_at, last_queried_at, actual_recorded_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(site_id, production_date) DO UPDATE SET
              base_forecast_kwh=excluded.base_forecast_kwh,
              rain_forecast_kwh=excluded.rain_forecast_kwh,
              base_display_kwh=excluded.base_display_kwh,
              rain_display_kwh=excluded.rain_display_kwh,
              source_type=excluded.source_type,
              actual_kwh=COALESCE(excluded.actual_kwh, solar_production_history.actual_kwh),
              actual_recorded_at=COALESCE(excluded.actual_recorded_at, solar_production_history.actual_recorded_at),
              last_queried_at=excluded.last_queried_at
            """,
            rows,
        )


def solar_history_for_month(site_id: str, month_start: date) -> Dict[str, Dict[str, object]]:
    month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1)
    with cache_connect() as conn:
        ensure_solar_history_schema(conn)
        conn.row_factory = sqlite3.Row
        rows = conn.execute(
            """
            SELECT production_date, base_display_kwh, rain_display_kwh,
                   source_type, last_queried_at
            FROM solar_production_history
            WHERE site_id=? AND production_date>=? AND production_date<?
            ORDER BY production_date
            """,
            (site_id, month_start.isoformat(), month_end.isoformat()),
        ).fetchall()
    return {str(row["production_date"]): dict(row) for row in rows}


def solar_calibration_factor(site_id: str, rain_effect: bool, today: date) -> Tuple[float, int]:
    if not site_id:
        return 1.0, 0

    forecast_column = "rain_forecast_kwh" if rain_effect else "base_forecast_kwh"
    cutoff = today - timedelta(days=SOLAR_CALIBRATION_LOOKBACK_DAYS)
    with cache_connect() as conn:
        ensure_solar_history_schema(conn)
        rows = conn.execute(
            f"""
            SELECT actual_kwh, {forecast_column}
            FROM solar_production_history
            WHERE site_id=?
              AND production_date>=?
              AND production_date<?
              AND actual_kwh IS NOT NULL
              AND {forecast_column}>0
            ORDER BY production_date DESC
            """,
            (site_id, cutoff.isoformat(), today.isoformat()),
        ).fetchall()

    ratios = [actual / forecast for actual, forecast in rows if actual >= 0 and forecast > 0]
    if len(ratios) < SOLAR_CALIBRATION_MIN_DAYS:
        return 1.0, len(ratios)

    factor = statistics.median(ratios)
    return min(max(factor, 0.35), 1.40), len(ratios)


# ---------------- HTML RENDER ----------------

def render_html(
        target_date: date,
        min_date: date,
        max_date: date,
        tz: str,
        model: str,
        places: List[Place],
        lists_info: List[Dict[str, str]],
        active_list_id: str,
        time_index: List[datetime],
        cell_map: Dict[str, Dict[str, Tuple[str, float, int, float]]],  # icon, precip_mm, pop_pct, solar_wm2
        from_cache: bool,
        nocache: bool,
        base_params: Dict[str, str],
        view: str,
    ) -> str:
    """
    Updated pill/badge colors (POP):
      - < 30%     -> white
      - 30–50%    -> light green
      - 51–80%    -> yellow
      - > 80%     -> red
    """

    is_solar_view = view == "solar"
    title = "Mindoro Solar Forecast" if is_solar_view else "Mindoro Rain Forecast"
    table_id = "solarTable" if is_solar_view else "rainTable"
    download_prefix = "solar-matrix" if is_solar_view else "rain-matrix"

    prev_date = target_date - timedelta(days=1)
    next_date = target_date + timedelta(days=1)
    prev_ok = prev_date >= min_date
    next_ok = next_date <= max_date

    date_buttons = []
    d = min_date
    while d <= max_date:
        active = "active" if d == target_date else ""
        href = build_url(base_params, d)
        label = d.strftime("%a %b %d")
        date_buttons.append(f"<a class='dchip {active}' href='{href}'>{label}</a>")
        d += timedelta(days=1)

    list_buttons = []
    for l_info in lists_info:
        active = "active" if l_info["id"] == active_list_id else ""

        # Build URL for switching list but maintaining date
        l_params = dict(base_params)
        l_params["list"] = l_info["id"]
        # keep the same target_date
        l_href = build_url(l_params, target_date)

        list_buttons.append(f"<a class='dchip {active}' href='{l_href}'>{l_info['name']}</a>")

    view_buttons = []
    for view_id, label in [("rain", "Rain"), ("solar", "Solar")]:
        active = "active" if view_id == view else ""
        v_params = dict(base_params)
        if view_id == "solar":
            v_params["view"] = "solar"
        else:
            v_params.pop("view", None)
        v_href = build_url(v_params, target_date)
        view_buttons.append(f"<a class='dchip {active}' href='{v_href}'>{label}</a>")

    forecast_label = target_date.strftime("%d-%b-%y").upper()

    # ----- Legend samples (cell background gradient for precipitation) -----
    precip_samples = [
        (0.0, "0.0 mm - walang ulan"),
        (1.0, "1.0 mm - ambon lang"),
        (2.5, "2.5 mm - ulan na"),
        (5.0, "5.0 mm - malakas na ulan"),
        (SCALE_MAX_MM, f"{SCALE_MAX_MM:.1f} mm+ - buhos na ulan"),
    ]
    precip_legend_items = "".join(
        f"""
        <div class="legend-item">
          <span class="legend-rect" style="background:{precip_bg_color(mm)}"></span>
          <span class="legend-text">{label}</span>
        </div>
        """
        for mm, label in precip_samples
    )

    # Updated POP (pill/badge) colors + legend
    POP_WHITE = "#FFFFFF"  # < 30
    POP_GREEN = "#D1E7DD"  # 30–50 (light green)
    POP_YELLOW = "#FFF3CD" # 51–80
    POP_RED = "#F8D7DA"    # > 80

    pill_legend_items = f"""
      <div class="legend-item">
        <span class="legend-pill" style="background:{POP_WHITE}">
          <span class="legend-pill-dot"></span>
          POP &lt; 30%
        </span>malabong umulan
      </div>
      <div class="legend-item">
        <span class="legend-pill" style="background:{POP_GREEN}">
          <span class="legend-pill-dot"></span>
          POP 30–50%
        </span>baka umulan
      </div>
      <div class="legend-item">
        <span class="legend-pill" style="background:{POP_YELLOW}">
          <span class="legend-pill-dot"></span>
          POP 51–80%
        </span>maghanda sa posibleng ulan
      </div>
      <div class="legend-item">
        <span class="legend-pill" style="background:{POP_RED}">
          <span class="legend-pill-dot"></span>
          POP &gt; 80%
        </span>asahang uulan
      </div>
    """

    solar_samples = [
        (0.0, "0 W/m² - no usable sun"),
        (250.0, "250 W/m² - weak sun"),
        (500.0, "500 W/m² - moderate sun"),
        (750.0, "750 W/m² - strong sun"),
        (SOLAR_SCALE_MAX_WM2, f"{SOLAR_SCALE_MAX_WM2:.0f} W/m² - excellent sun"),
    ]
    solar_legend_items = "".join(
        f"""
        <div class="legend-item">
          <span class="legend-rect" style="background:{solar_bg_color(wm2)}"></span>
          <span class="legend-text">{label}</span>
        </div>
        """
        for wm2, label in solar_samples
    )

    if is_solar_view:
        legend_html = f"""
<div class="legend" aria-label="Legend">
  <h3>Legend</h3>
  <div class="legend-grid">
    <div class="legend-block">
      <div style="font-size:13px;font-weight:900;margin-bottom:8px;">Solar Irradiance</div>
      <div class="legend-items">
        {solar_legend_items}
      </div>
    </div>
  </div>
</div>
"""
    else:
        legend_html = f"""
<div class="legend" aria-label="Legend">
  <h3>Legend</h3>
  <div class="legend-grid">
    <div class="legend-block">
      <div style="font-size:13px;font-weight:900;margin-bottom:8px;">Chance of Rain</div>
      <div class="legend-items">
        {pill_legend_items}
      </div>
    </div>

    <div class="legend-block">
      <div style="font-size:13px;font-weight:900;margin-bottom:8px;">Rain Intensity</div>
      <div class="legend-items">
        {precip_legend_items}
      </div>
    </div>
  </div>
</div>
"""

    html = f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>{title}</title>

<style>
body {{ font-family: Segoe UI, Arial, sans-serif; margin:16px; }}

.navbar {{
  display:flex;
  gap:12px;
  align-items:center;
  flex-wrap:wrap;
  margin: 6px 0 14px;
}}

.btn {{
  display:inline-block;
  padding:7px 14px;
  border-radius:10px;
  background:#111;
  color:#fff;
  text-decoration:none;
  font-weight:700;
  font-size:13px;
}}

.dchips {{
  display:flex;
  gap:8px;
  flex-wrap:wrap;
}}

.dchip {{
  display:inline-block;
  padding:6px 10px;
  border-radius:999px;
  background:#fff;
  border:1px solid #ddd;
  text-decoration:none;
  color:#111;
  font-weight:700;
  font-size:13px;
}}

.dchip.active {{
  border-color:#111;
  box-shadow:0 1px 2px rgba(0,0,0,15);
}}

/* --- MOBILE-FRIENDLY TABLE: allow horizontal scroll instead of compressing --- */
.table-wrap {{
  overflow-x:auto;
  -webkit-overflow-scrolling:touch;
  border:1px solid #ddd;
  border-radius:12px;
  background:#fff;
}}

table {{
  width:max-content;
  min-width:100%;
  table-layout:fixed;
  border-collapse:separate;
  border-spacing:0;
  background:#ffffff;
}}

th, td {{ border:1px solid #ddd; padding:8px; text-align:center; }}

th {{
  background:#f5f5f5;
  position:sticky;
  top:0;
  z-index:3;
}}

th.timehead {{
  left:0;
  z-index:4;
  position:sticky;
}}

td.time {{
  width:90px;
  position:sticky;
  left:0;
  z-index:2;
  font-weight:700;
}}

.time-pill {{
  display:inline-block;
  padding:4px 10px;
  border-radius:999px;
  background:#fff;
  font-size:13px;
  font-weight:800;
  color:#111;
  box-shadow:0 1px 2px rgba(0,0,0,20);
}}

.pill {{
  display:inline-block;
  padding:4px 8px;
  background:#fff;
  border-radius:999px;
  box-shadow:0 1px 2px rgba(0,0,0,18);
  color:#111;
  line-height:1;
  white-space:nowrap;
}}

.pill .icon {{
  display:inline-block;
  vertical-align:middle;
  font-size:16px;
  margin-right:6px;
}}

.pill .val {{
  display:inline-block;
  vertical-align:middle;
  font-size:13px;
  font-weight:700;
  color:#111;
}}

.pill .pop {{
  display:inline-block;
  vertical-align:middle;
  font-size:11px;
  font-weight:800;
  margin-left:6px;
  padding:2px 6px;
  border-radius:999px;
  background:rgba(0,0,0,0.06);
  color:#111;
}}

.solar-val {{
  display:inline-block;
  min-width:56px;
  padding:4px 8px;
  border-radius:999px;
  background:rgba(255,255,255,0.86);
  box-shadow:0 1px 2px rgba(0,0,0,18);
  color:#111;
  font-size:13px;
  font-weight:900;
  line-height:1;
}}

.total-label {{
  background:#111827 !important;
  color:#fff;
}}

.total-pill {{
  display:inline-block;
  min-width:64px;
  padding:5px 10px;
  border-radius:999px;
  background:#111827;
  color:#fff;
  box-shadow:0 1px 2px rgba(0,0,0,22);
  font-size:13px;
  font-weight:900;
  line-height:1;
}}

/* Give each place column a minimum width so it stays readable */
th:not(.timehead), td:not(.time) {{
  min-width:92px;
}}

@media (max-width: 520px) {{
  body {{ margin:12px; }}
  th, td {{ padding:6px; }}
  th:not(.timehead), td:not(.time) {{ min-width:86px; }}
  .pill {{ padding:3px 7px; }}
  .pill .icon {{ font-size:14px; margin-right:5px; }}
  .pill .val {{ font-size:12px; }}
  .time-pill {{ font-size:12px; padding:3px 9px; }}
  .pill .pop {{ font-size:10px; padding:2px 5px; }}
  .solar-val {{ font-size:12px; min-width:50px; padding:3px 7px; }}
  .total-pill {{ font-size:12px; min-width:56px; padding:4px 8px; }}
}}

/* ---------- LEGEND ---------- */
.legend {{
  margin-top:14px;
  padding:12px;
  border:1px solid #ddd;
  border-radius:12px;
  background:#fff;
}}

.legend h3 {{
  margin:0 0 8px;
  font-size:14px;
}}

.legend-grid {{
  display:flex;
  flex-wrap:wrap;
  gap:18px;
}}

.legend-block {{
  min-width:220px;
}}

.legend-items {{
  display:flex;
  flex-direction:column;
  gap:8px;
}}

.legend-item {{
  display:flex;
  align-items:center;
  gap:10px;
}}

.legend-rect {{
  width:34px;
  height:18px;
  border-radius:4px;
  border:1px solid rgba(0,0,0,0.18);
  box-shadow:0 1px 1px rgba(0,0,0,0.08);
}}

.legend-text {{
  font-size:13px;
  font-weight:700;
  color:#111;
}}

.legend-pill {{
  display:inline-flex;
  align-items:center;
  gap:8px;
  padding:5px 10px;
  border-radius:999px;
  border:1px solid rgba(0,0,0,0.10);
  box-shadow:0 1px 2px rgba(0,0,0,0.12);
  font-size:13px;
  font-weight:800;
  color:#111;
}}

.legend-pill-dot {{
  width:10px;
  height:10px;
  border-radius:999px;
  background:rgba(0,0,0,0.18);
}}
</style>
</head>

<body>

<h2 style="margin:0 0 6px;">{title}</h2>

<div class="navbar">
  <div class="dchips">
    {''.join(list_buttons)}
  </div>
</div>

<div class="navbar">
  <div class="dchips">
    {''.join(date_buttons)}
  </div>

  <div class="dchips">
    {''.join(view_buttons)}
  </div>

  <a href="/solar-site" class="btn">Site Production</a>
  <a href="#" class="btn" id="downloadBtn">Download JPG</a>
</div>

<div class="table-wrap">
  <table id="{table_id}">
    <tr>
      <th class="timehead">{forecast_label}</th>
      {''.join(f"<th>{p.label}</th>" for p in places)}
    </tr>
"""

    for t in time_index:
        hk = t.strftime("%I:00 %p")
        tbg = time_bg_color(t)

        html += (
            f"<tr>"
            f"<td class='time' style='background:{tbg}'>"
            f"<span class='time-pill'>{hk}</span>"
            f"</td>"
        )

        for p in places:
            icon, precip, pop, solar = cell_map.get(p.label, {}).get(
                t.strftime("%H:00"), ("—", 0.0, 0, 0.0)
            )

            if is_solar_view:
                html += (
                    f"<td style='background:{solar_bg_color(solar)}'>"
                    f"<span class='solar-val'>{solar_display(solar)}</span>"
                    f"</td>"
                )
            else:
                bg = precip_bg_color(precip)
                val = precip_display(precip)

                # Updated badge/pill color based on precipitation probability
                pop_i = int(pop or 0)
                if pop_i > 80:
                    pill_bg = POP_RED       # > 80
                elif pop_i > 50:
                    pill_bg = POP_YELLOW    # 51–80
                elif pop_i >= 30:
                    pill_bg = POP_GREEN     # 30–50
                else:
                    pill_bg = POP_WHITE     # < 30

                html += (
                    f"<td style='background:{bg}'>"
                    f"<span class='pill' style='background:{pill_bg}'>"
                    f"<span class='icon'>{icon}</span>"
                    f"<span class='val'>{val}</span>"
                    f"</span>"
                    f"</td>"
                )

        html += "</tr>"

    if is_solar_view:
        html += (
            "<tr>"
            "<td class='time total-label'>"
            "<span class='time-pill'>TOTAL</span>"
            "</td>"
        )


        for p in places:
            total_solar = sum(
                cell_map.get(p.label, {}).get(t.strftime("%H:00"), ("—", 0.0, 0, 0.0))[3]
                for t in time_index
            )
            html += (
                "<td class='total-cell'>"
                f"<span class='total-pill'>{solar_total_display(total_solar)}</span>"
                "</td>"
            )

        html += "</tr>"

    html += f"""
  </table>
</div>

<!-- LEGEND (after table) -->
{legend_html}

<script src="https://cdn.jsdelivr.net/npm/html2canvas@1.4.1/dist/html2canvas.min.js"></script>
<script>
document.getElementById("downloadBtn").addEventListener("click", function (e) {{
  e.preventDefault();

  const table = document.getElementById("{table_id}");

  html2canvas(table, {{
    backgroundColor: "#ffffff",
    scale: 2,
    useCORS: true
  }}).then(canvas => {{
    const link = document.createElement("a");
    const d = new Date().toISOString().slice(0,10);

    link.download = `{download_prefix}-${{d}}.jpg`;
    link.href = canvas.toDataURL("image/jpeg", 0.95);
    link.click();
  }});
}});
</script>

</body>
</html>
"""
    return html


def render_solar_site_html(
        site_name: str,
        site_id: str,
        sites: List[Dict[str, object]],
        tz: str,
        model: str,
        daily_rows: List[Dict[str, object]],
        battery_days: List[Dict[str, object]],
        hourly_days: List[Dict[str, object]],
        battery_capacity_ah: float,
        max_charging_ampere: float,
        initial_soc_percent: float,
        stop_discharge_soc_percent: float,
        consumption_kwh_per_hour: float,
        charge_first: bool,
        rain_effect: bool,
        calibration_factor: float,
        calibration_days: int,
        history_month: date,
        history_rows: Dict[str, Dict[str, object]],
        inverter_note: str,
        today: date,
    ) -> str:
    safe_name = html_lib.escape(site_name)
    max_kwh = max((_to_float(row["production_kwh"]) for row in daily_rows), default=0.0)
    axis_step = 25
    axis_max = nice_solar_axis_max(max_kwh, axis_step)
    axis_ticks = list(range(axis_max, -1, -axis_step))

    site_options = ""
    for site in sites:
        option_id = str(site.get("id", "")).strip()
        option_name = str(site.get("name", option_id)).strip() or option_id
        selected = " selected" if option_id == site_id else ""
        site_options += (
            f'<option value="{html_lib.escape(option_id)}"{selected}>'
            f'{html_lib.escape(option_name)}</option>'
        )

    if not site_options:
        site_options = '<option value="">No saved sites</option>'

    y_axis_html = '<div class="y-unit">kWh</div>'
    grid_lines_html = ""
    for tick in axis_ticks:
        tick_pct = (tick / axis_max) * 100.0 if axis_max else 0.0
        tick_top = (1.0 - (tick_pct / 100.0)) * 220.0
        grid_top = 26.0 + tick_top
        y_axis_html += f'<span class="y-tick" style="top:{tick_top:.2f}px;">{tick:,}</span>'
        grid_lines_html += f'<span class="grid-line" style="top:{grid_top:.2f}px;"></span>'

    bars_html = ""
    for day_index, row in enumerate(daily_rows):
        day = row["date"]
        if isinstance(day, date):
            date_label = day.strftime("%b%d")
            weekday_label = day.strftime("%a")
        else:
            date_label = html_lib.escape(str(day))
            weekday_label = ""

        production_kwh = _to_float(row["production_kwh"])
        base_forecast_kwh = _to_float(row.get("base_display_kwh"))
        rain_forecast_kwh = _to_float(row.get("rain_display_kwh"))
        bar_height = (production_kwh / axis_max) * 100.0 if axis_max else 0.0
        bg = production_bar_color(production_kwh, max_kwh)
        value_label = f"{int(round(production_kwh)):,}"
        conservative_kwh = _to_float(row.get("conservative_display_kwh"), rain_forecast_kwh)
        upper_kwh = _to_float(row.get("upper_display_kwh"), rain_forecast_kwh)
        comparison_label = (
            f"Range {int(round(conservative_kwh)):,}-{int(round(upper_kwh)):,} kWh"
            if rain_effect
            else f"Rain-aware {int(round(rain_forecast_kwh)):,} kWh"
        )
        bars_html += f"""
        <button type="button" class="bar-slot day-bar" data-hourly-index="{day_index}"
                aria-controls="hourlyForecastCard" aria-pressed="{'true' if day_index == 0 else 'false'}"
                title="View {date_label} hourly forecast">
          <div class="bar-value">{value_label}</div>
          <div class="bar-track" aria-label="{date_label} estimated production {value_label} kWh">
            <div class="bar-fill" style="height:{bar_height:.2f}%;background:{bg};"></div>
          </div>
          <div class="x-label">
            <span>{date_label}</span>
            <strong>{weekday_label}</strong>
            <small>{comparison_label}</small>
          </div>
        </button>
"""

    mode_name = "Smart rain v7" if rain_effect else "Standard v1"
    rain_model_count = max(
        (int(_to_float(row.get("rain_model_count"), 1.0)) for row in daily_rows),
        default=1,
    )
    method_note = (
        f"Smart model chosen each hour from {rain_model_count} weather models"
        if rain_effect
        else "Rain effect is off"
    )
    calibration_note = (
        f"Site calibration: {calibration_factor:.2f}x from {calibration_days} measured days"
        if calibration_days >= SOLAR_CALIBRATION_MIN_DAYS
        else (
            f"Site calibration: learning ({calibration_days}/{SOLAR_CALIBRATION_MIN_DAYS} completed inverter days)"
            if inverter_note
            else "Site calibration: default (no measured production feed)"
        )
    )
    inverter_status_html = (
        f"<span>{html_lib.escape(inverter_note)}</span>"
        if inverter_note
        else ""
    )
    today_row = next((row for row in daily_rows if row.get("date") == today), None)
    today_progress_html = ""
    if today_row:
        by_now_kwh = max(0.0, _to_float(today_row.get("estimated_by_now_kwh")))
        remaining_kwh = max(0.0, _to_float(today_row.get("remaining_today_kwh")))
        full_day_kwh = max(0.0, _to_float(today_row.get("production_kwh")))
        by_now_label = (
            "Inverter actual by now"
            if _to_float(today_row.get("has_inverter_actual")) > 0
            else "Estimated by now"
        )
        conservative_kwh = max(0.0, _to_float(today_row.get("conservative_display_kwh")))
        upper_kwh = max(0.0, _to_float(today_row.get("upper_display_kwh")))
        consensus_note = (
            f"Expected range: {conservative_kwh:,.1f}-{upper_kwh:,.1f} kWh"
            if rain_effect
            else mode_name
        )
        today_progress_html = f"""
        <div class="today-progress">
          <span><small>{by_now_label}</small><strong>{by_now_kwh:,.1f} kWh</strong></span>
          <span><small>Remaining today</small><strong>{remaining_kwh:,.1f} kWh</strong></span>
          <span><small>Full-day estimate</small><strong>{full_day_kwh:,.1f} kWh</strong></span>
          <em>{consensus_note}</em>
        </div>
"""

    history_forecast_key = "rain_display_kwh" if rain_effect else "base_display_kwh"
    history_cells = ""
    history_values: List[float] = []
    month_calendar = calendar_lib.Calendar(firstweekday=0)
    for calendar_day in [day for week in month_calendar.monthdatescalendar(history_month.year, history_month.month) for day in week]:
        if calendar_day.month != history_month.month:
            history_cells += '<div class="calendar-day outside" aria-hidden="true"></div>'
            continue

        history = history_rows.get(calendar_day.isoformat(), {})
        production = history.get(history_forecast_key)
        source_type = str(history.get("source_type") or "forecast").lower()
        source_label = {
            "actual": "Inverter actual",
            "actual + forecast": "Actual + forecast",
            "nowcast": "Nowcast",
        }.get(source_type, "Forecast")
        is_today = " today" if calendar_day == today else ""
        if production is not None:
            production_value = max(0.0, _to_float(production))
            history_values.append(production_value)
            production_html = f'<strong>{production_value:,.1f}<span> kWh</span></strong>'
            source_html = f'<small>{source_label} &middot; {mode_name}</small>'
            cell_state = " has-production"
        else:
            production_html = '<em>No saved query</em>'
            source_html = ""
            cell_state = ""
        history_cells += f"""
        <div class="calendar-day{cell_state}{is_today}">
          <span class="day-number">{calendar_day.day}</span>
          {production_html}
          {source_html}
        </div>
"""

    previous_month = (history_month - timedelta(days=1)).replace(day=1)
    next_month = (history_month.replace(day=28) + timedelta(days=4)).replace(day=1)

    def history_url(month: date) -> str:
        params = {
            "site": site_id,
            "model": model,
            "tz": tz,
            "history_month": month.strftime("%Y-%m"),
            "rain": "1" if rain_effect else "0",
        }
        return "/solar-site?" + urlencode(params)

    history_total = sum(history_values)
    history_summary = (
        f"{len(history_values)} saved query days &middot; {history_total:,.1f} kWh total &middot; "
        f"{history_total / len(history_values):,.1f} kWh/day"
        if history_values
        else "No forecast or nowcast query has been saved for this month."
    )

    site_name_json = json.dumps(site_name)
    battery_forecast_json = json.dumps(
        {
            "capacityAh": max(0.0, battery_capacity_ah),
            "maxChargingAmpere": max(0.0, max_charging_ampere),
            "nominalVoltage": BATTERY_NOMINAL_VOLTAGE,
            "initialSocPercent": min(max(initial_soc_percent, 0.0), 100.0),
            "stopDischargeSocPercent": min(max(stop_discharge_soc_percent, 0.0), 100.0),
            "consumptionKwhPerHour": max(0.0, consumption_kwh_per_hour),
            "chargeFirst": bool(charge_first),
            "rainEffect": bool(rain_effect),
            "days": battery_days,
        }
    )
    hourly_forecast_json = json.dumps({"days": hourly_days})

    return f"""<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>Solar Production Forecast - {safe_name}</title>
<style>
body {{
  font-family: Segoe UI, Arial, sans-serif;
  margin:0;
  background:#F8FAFC;
  color:#111827;
}}

.page {{
  max-width:980px;
  margin:0 auto;
  padding:18px;
}}

.topbar {{
  display:flex;
  justify-content:space-between;
  align-items:center;
  gap:12px;
  flex-wrap:wrap;
  margin-bottom:14px;
}}

h1 {{
  font-size:24px;
  margin:0;
  letter-spacing:0;
}}

.navlink {{
  display:inline-block;
  padding:7px 12px;
  border:1px solid #CBD5E1;
  border-radius:8px;
  color:#111827;
  background:#fff;
  text-decoration:none;
  font-size:13px;
  font-weight:800;
}}

.top-actions {{
  display:flex;
  gap:8px;
  flex-wrap:wrap;
}}

.site-panel {{
  background:#fff;
  border:1px solid #CBD5E1;
  border-radius:8px;
  padding:14px;
  margin-bottom:14px;
}}

.site-select-row {{
  display:flex;
  gap:10px;
  align-items:end;
  flex-wrap:wrap;
}}

.site-select-row .select-wrap {{
  flex:1 1 260px;
}}

.mode-control {{
  display:flex;
  align-items:center;
  gap:8px;
  min-height:37px;
  margin:0;
  padding:0 4px;
  color:#0F172A;
  font-size:13px;
}}

.mode-control input {{
  width:17px;
  height:17px;
  accent-color:#0369A1;
}}

.mode-status {{
  display:flex;
  justify-content:space-between;
  gap:10px;
  flex-wrap:wrap;
  margin-top:10px;
  padding-top:10px;
  border-top:1px solid #E2E8F0;
  color:#475569;
  font-size:12px;
  font-weight:700;
}}

.mode-status strong {{ color:#0F172A; }}

label {{
  display:block;
  font-size:12px;
  font-weight:800;
  color:#334155;
  margin-bottom:4px;
}}

select {{
  width:100%;
  box-sizing:border-box;
  padding:8px 9px;
  border:1px solid #CBD5E1;
  border-radius:6px;
  font-size:14px;
  background:#fff;
}}

button {{
  width:100%;
  padding:9px 12px;
  border:0;
  border-radius:6px;
  background:#111827;
  color:#fff;
  font-size:14px;
  font-weight:900;
  cursor:pointer;
}}

.forecast-export {{
  background:#fff;
  border:1px solid #CBD5E1;
  border-radius:8px;
  padding:14px;
}}

.export-heading {{
  display:flex;
  justify-content:space-between;
  gap:12px;
  align-items:flex-end;
  flex-wrap:wrap;
  margin-bottom:12px;
}}

.export-actions {{
  display:flex;
  align-items:flex-end;
  justify-content:flex-end;
  gap:8px;
  flex-wrap:wrap;
}}

.card-export-btn {{
  width:auto;
  min-height:35px;
  display:inline-flex;
  align-items:center;
  justify-content:center;
  gap:7px;
  padding:8px 11px;
  white-space:nowrap;
}}

.card-export-btn:disabled {{
  opacity:.6;
  cursor:wait;
}}

.jpeg-exporting .card-export-btn {{
  display:none;
}}

.export-download-icon {{
  font-size:16px;
  line-height:1;
}}

.export-kicker {{
  font-size:12px;
  font-weight:900;
  color:#475569;
  text-transform:uppercase;
  letter-spacing:0;
}}

.export-title {{
  font-size:24px;
  font-weight:900;
  margin-top:2px;
}}

.today-progress {{
  display:grid;
  grid-template-columns:repeat(3, minmax(130px, 1fr)) auto;
  gap:12px;
  align-items:center;
  margin-bottom:14px;
  padding:10px 0;
  border-top:1px solid #E2E8F0;
  border-bottom:1px solid #E2E8F0;
}}

.today-progress span {{ display:grid; gap:2px; }}
.today-progress small {{ color:#64748B; font-size:10px; font-weight:800; }}
.today-progress strong {{ color:#0F172A; font-size:16px; }}
.today-progress em {{ color:#0369A1; font-size:11px; font-style:normal; font-weight:900; text-align:right; }}

.chart-card {{
  display:grid;
  grid-template-columns:58px minmax(0, 1fr);
  gap:12px;
  min-height:350px;
}}

.y-axis {{
  position:relative;
  height:220px;
  margin-top:26px;
  margin-bottom:44px;
  color:#475569;
  font-size:12px;
  font-weight:900;
}}

.y-unit {{
  position:absolute;
  right:0;
  bottom:calc(100% + 10px);
  color:#334155;
  font-size:12px;
  font-weight:900;
}}

.y-tick {{
  position:absolute;
  right:0;
  transform:translateY(-50%);
  line-height:1;
  white-space:nowrap;
}}

.chart-body {{
  position:relative;
  display:grid;
  grid-template-columns:repeat(7, minmax(54px, 1fr));
  gap:12px;
  min-width:620px;
}}

.grid-line {{
  position:absolute;
  left:0;
  right:0;
  height:0;
  border-top:1px dotted #CBD5E1;
  pointer-events:none;
}}

.bar-slot {{
  position:relative;
  z-index:1;
  display:grid;
  grid-template-rows:20px minmax(220px, 1fr) 58px;
  gap:6px;
  min-width:0;
}}

.day-bar {{
  appearance:none;
  width:auto;
  padding:0;
  border:0;
  border-radius:0;
  background:transparent;
  color:inherit;
  font:inherit;
  cursor:pointer;
}}

.day-bar:hover {{ background:#F8FAFC; }}

.day-bar:focus-visible {{
  outline:2px solid #0284C7;
  outline-offset:2px;
}}

.day-bar[aria-pressed="true"] .bar-track {{
  border-bottom-color:#0284C7;
  box-shadow:inset 0 -3px 0 #0284C7;
}}

.day-bar[aria-pressed="true"] .x-label span {{ color:#0369A1; }}

.bar-value {{
  text-align:center;
  font-size:12px;
  font-weight:900;
  color:#111827;
}}

.bar-track {{
  display:flex;
  align-items:flex-end;
  height:100%;
  min-height:220px;
  border-bottom:1px solid #94A3B8;
}}

.bar-fill {{
  width:100%;
  min-height:2px;
  border-radius:6px 6px 0 0;
  box-shadow:inset 0 1px 0 rgba(255,255,255,.45);
}}

.x-label {{
  text-align:center;
  font-size:12px;
  font-weight:900;
  color:#111827;
  line-height:1.05;
}}

.x-label strong {{
  display:block;
  margin-top:4px;
  font-size:11px;
  color:#64748B;
}}

.x-label small {{
  display:block;
  margin-top:4px;
  color:#64748B;
  font-size:9px;
  font-weight:700;
  line-height:1.15;
}}

.chart-scroll {{
  overflow-x:auto;
  -webkit-overflow-scrolling:touch;
}}

.chart-legend {{
  display:flex;
  justify-content:flex-end;
  gap:10px;
  flex-wrap:wrap;
  margin-top:12px;
  color:#475569;
  font-size:12px;
  font-weight:800;
}}

.legend-item {{
  display:inline-flex;
  align-items:center;
  gap:5px;
}}

.legend-swatch {{
  width:12px;
  height:12px;
  border-radius:3px;
}}

.hourly-card {{
  margin-top:14px;
  scroll-margin-top:14px;
}}

.hourly-toolbar {{
  width:min(240px, 100%);
}}

.hourly-table-wrap {{
  overflow-x:auto;
  -webkit-overflow-scrolling:touch;
}}

.hourly-table {{
  width:100%;
  min-width:850px;
  border-collapse:collapse;
  font-size:12px;
}}

.hourly-table th,
.hourly-table td {{
  padding:9px 10px;
  border-bottom:1px solid #E2E8F0;
  text-align:left;
  vertical-align:top;
}}

.hourly-table th {{
  color:#475569;
  font-size:10px;
  text-transform:uppercase;
  background:#F8FAFC;
}}

.hourly-table td:first-child,
.hourly-table td:nth-child(2) {{
  color:#0F172A;
  font-weight:900;
  white-space:nowrap;
}}

.hourly-table td:nth-child(3),
.hourly-table td:nth-child(4),
.hourly-table td:nth-child(5),
.hourly-table td:nth-child(6) {{
  white-space:nowrap;
}}

.rain-signal {{
  display:inline-flex;
  align-items:center;
  gap:6px;
  font-weight:800;
}}

.rain-signal::before {{
  content:"";
  width:8px;
  height:8px;
  border-radius:50%;
  background:#94A3B8;
  flex:0 0 auto;
}}

.rain-signal[data-level="dry"]::before {{ background:#16A34A; }}
.rain-signal[data-level="light"]::before {{ background:#EAB308; }}
.rain-signal[data-level="likely"]::before {{ background:#EA580C; }}
.rain-signal[data-level="heavy"]::before {{ background:#DC2626; }}
.rain-signal[data-level="actual"]::before {{ background:#0284C7; }}
.rain-signal[data-level="off"]::before {{ background:#64748B; }}

.hourly-interpretation {{
  min-width:250px;
  color:#475569;
  line-height:1.35;
}}

.selection-rules {{
  margin-top:14px;
  padding-top:12px;
  border-top:1px solid #CBD5E1;
}}

.selection-rules h3 {{
  margin:0 0 8px;
  color:#0F172A;
  font-size:14px;
}}

.selection-rule {{
  display:grid;
  grid-template-columns:10px minmax(110px, .7fr) minmax(210px, 1.3fr) minmax(220px, 1.4fr);
  gap:8px;
  align-items:center;
  padding:7px 0;
  border-bottom:1px solid #F1F5F9;
  color:#475569;
  font-size:11px;
}}

.selection-rule::before {{
  content:"";
  width:8px;
  height:8px;
  border-radius:50%;
  background:#94A3B8;
}}

.selection-rule[data-level="dry"]::before {{ background:#16A34A; }}
.selection-rule[data-level="light"]::before {{ background:#EAB308; }}
.selection-rule[data-level="likely"]::before {{ background:#EA580C; }}
.selection-rule[data-level="heavy"]::before {{ background:#DC2626; }}
.selection-rule[data-level="observed"]::before {{ background:#0284C7; }}
.selection-rule[data-level="actual"]::before {{ background:#0F766E; }}
.selection-rule strong {{ color:#0F172A; }}
.selection-rule em {{ color:#334155; font-style:normal; font-weight:800; }}

.selection-rules p {{
  margin:9px 0 0;
  color:#64748B;
  font-size:11px;
}}

.battery-card {{
  margin-top:14px;
}}

.battery-controls {{
  display:grid;
  grid-template-columns:1.2fr 1fr 1fr;
  gap:10px;
  align-items:end;
  margin-bottom:12px;
}}

.battery-control input {{
  width:100%;
  box-sizing:border-box;
  padding:8px 9px;
  border:1px solid #CBD5E1;
  border-radius:6px;
  font-size:14px;
  background:#fff;
}}

.battery-check {{
  display:flex;
  align-items:center;
  gap:8px;
  min-height:37px;
  font-size:13px;
  font-weight:900;
  color:#334155;
}}

.battery-check input {{
  width:16px;
  height:16px;
}}

.battery-series-controls {{
  display:flex;
  align-items:center;
  justify-content:flex-end;
  gap:16px;
  flex-wrap:wrap;
  margin:2px 0 8px;
}}

.battery-series-toggle {{
  display:flex;
  align-items:center;
  gap:7px;
  color:#334155;
  font-size:12px;
  font-weight:900;
  cursor:pointer;
}}

.battery-series-toggle input {{
  width:16px;
  height:16px;
}}

.battery-series-swatch {{
  width:18px;
  height:3px;
  border-radius:2px;
  background:var(--series-color);
}}

.battery-chart-wrap {{
  overflow-x:auto;
  -webkit-overflow-scrolling:touch;
}}

.battery-chart {{
  display:block;
  width:100%;
  min-width:620px;
  height:280px;
}}

.battery-summary {{
  display:flex;
  gap:10px;
  justify-content:space-between;
  flex-wrap:wrap;
  margin-top:10px;
  color:#475569;
  font-size:13px;
  font-weight:800;
}}

.battery-summary strong {{
  color:#111827;
  font-size:16px;
}}

.battery-chart circle {{
  cursor:pointer;
}}

.history-card {{
  margin-top:14px;
}}

.history-heading {{
  display:flex;
  justify-content:space-between;
  align-items:center;
  gap:12px;
  flex-wrap:wrap;
  margin-bottom:12px;
}}

.month-nav {{
  display:grid;
  grid-template-columns:36px minmax(150px, 1fr) 36px;
  align-items:center;
  gap:6px;
}}

.month-nav a {{
  display:grid;
  place-items:center;
  height:34px;
  border:1px solid #CBD5E1;
  border-radius:6px;
  background:#fff;
  color:#0F172A;
  text-decoration:none;
  font-size:20px;
  line-height:1;
}}

.month-nav strong {{
  text-align:center;
  font-size:14px;
}}

.calendar-scroll {{
  overflow-x:auto;
  -webkit-overflow-scrolling:touch;
}}

.calendar-grid {{
  display:grid;
  grid-template-columns:repeat(7, minmax(94px, 1fr));
  min-width:700px;
  border-top:1px solid #CBD5E1;
  border-left:1px solid #CBD5E1;
}}

.weekday {{
  padding:7px 4px;
  border-right:1px solid #CBD5E1;
  border-bottom:1px solid #CBD5E1;
  background:#F1F5F9;
  color:#475569;
  text-align:center;
  font-size:11px;
  font-weight:900;
}}

.calendar-day {{
  position:relative;
  display:flex;
  min-height:78px;
  padding:8px;
  box-sizing:border-box;
  flex-direction:column;
  justify-content:flex-end;
  gap:4px;
  border-right:1px solid #CBD5E1;
  border-bottom:1px solid #CBD5E1;
  background:#fff;
}}

.calendar-day.outside {{ background:#F8FAFC; }}
.calendar-day.has-production {{ background:#ECFDF5; }}
.calendar-day.today {{ box-shadow:inset 0 0 0 2px #0284C7; }}

.day-number {{
  position:absolute;
  top:7px;
  right:8px;
  color:#475569;
  font-size:11px;
  font-weight:900;
}}

.calendar-day strong {{
  color:#065F46;
  font-size:15px;
}}

.calendar-day strong span {{ font-size:9px; }}

.calendar-day em {{
  color:#94A3B8;
  font-size:10px;
  font-style:normal;
}}

.calendar-day small {{
  color:#475569;
  font-size:9px;
  line-height:1.15;
}}

.history-summary {{
  margin:10px 0 12px;
  color:#475569;
  font-size:12px;
  font-weight:800;
}}

@media (max-width: 760px) {{
  .page {{ padding:12px; }}
  .chart-card {{ grid-template-columns:52px minmax(0, 1fr); }}
  .chart-body {{ min-width:560px; gap:9px; }}
  .battery-controls {{ grid-template-columns:1fr; }}
  .today-progress {{ grid-template-columns:1fr 1fr; }}
  .today-progress em {{ text-align:left; }}
  .selection-rule {{ grid-template-columns:10px 1fr; align-items:start; }}
  .selection-rule::before {{ grid-row:1 / span 3; margin-top:3px; }}
}}

@media (max-width: 460px) {{
  h1 {{ font-size:21px; }}
}}
</style>
</head>
<body>
<main class="page">
  <div class="topbar">
    <h1>Solar Production Forecast</h1>
    <div class="top-actions">
      <a class="navlink" href="/?view=solar">Solar Matrix</a>
      <a class="navlink" href="#" id="siteDownloadBtn">Download JPG</a>
    </div>
  </div>

  <section class="site-panel">
    <form method="get" action="/solar-site" class="site-select-row">
      <div class="select-wrap">
        <label for="site">Saved Site</label>
        <select id="site" name="site" onchange="this.form.submit()">
          {site_options}
        </select>
      </div>
      <input type="hidden" name="model" value="{html_lib.escape(model)}" />
      <input type="hidden" name="tz" value="{html_lib.escape(tz)}" />
      <input type="hidden" name="history_month" value="{history_month.strftime('%Y-%m')}" />
      <input type="hidden" name="rain" value="0" />
      <label class="mode-control" for="rainEffect">
        <input id="rainEffect" name="rain" type="checkbox" value="1"{' checked' if rain_effect else ''} onchange="this.form.submit()" />
        Include rain effect
      </label>
      <div>
        <button type="submit">Show Site</button>
      </div>
    </form>
    <div class="mode-status">
      <span>Prediction: <strong>{mode_name}</strong></span>
      <span>{method_note}</span>
      <span>{calibration_note}</span>
      {inverter_status_html}
    </div>
  </section>

  <section class="forecast-export" id="solarSiteCapture">
    <div class="export-heading">
      <div>
        <div class="export-kicker">Solar Production Forecast &middot; {mode_name}</div>
        <div class="export-title">{safe_name}</div>
      </div>
    </div>
    {today_progress_html}

    <div class="chart-scroll" role="group" aria-label="Select a day to view its hourly solar forecast">
      <div class="chart-card">
        <div class="y-axis" aria-hidden="true">
          {y_axis_html}
        </div>
        <div>
          <div class="chart-body">
            {grid_lines_html}
{bars_html}
          </div>
        </div>
      </div>
    </div>

    <div class="chart-legend" aria-label="Production color scale">
      <span class="legend-item"><span class="legend-swatch" style="background:#D1D5DB"></span>Low</span>
      <span class="legend-item"><span class="legend-swatch" style="background:#FDE68A"></span>Moderate</span>
      <span class="legend-item"><span class="legend-swatch" style="background:#FBBF24"></span>Strong</span>
      <span class="legend-item"><span class="legend-swatch" style="background:#F97316"></span>Peak</span>
    </div>
  </section>

  <section class="forecast-export hourly-card" id="hourlyForecastCard">
    <div class="export-heading">
      <div>
        <div class="export-kicker">Hourly Prediction &amp; Interpretation</div>
        <div class="export-title">{safe_name}</div>
      </div>
      <div class="export-actions">
        <div class="hourly-toolbar">
          <label for="hourlyDate">Forecast Date</label>
          <select id="hourlyDate"></select>
        </div>
        <button type="button" class="card-export-btn" id="hourlyDownloadBtn" title="Save hourly prediction as JPEG">
          <span class="export-download-icon" aria-hidden="true">&#8595;</span>
          Save JPEG
        </button>
      </div>
    </div>
    <div class="hourly-table-wrap">
      <table class="hourly-table">
        <thead>
          <tr>
            <th>Hour</th>
            <th>Expected</th>
            <th>Model range</th>
            <th>Rain</th>
            <th>Extra rain effect</th>
            <th>Model used</th>
            <th>Interpretation</th>
          </tr>
        </thead>
        <tbody id="hourlyRows"></tbody>
      </table>
    </div>
    <div class="selection-rules" aria-labelledby="selectionRulesTitle">
      <h3 id="selectionRulesTitle">Conditions Used For Model Selection</h3>
      <div class="selection-rule" data-level="dry">
        <strong>Dry</strong><span>Below 30% rain probability or below 0.05 mm</span><em>Middle raw solar model; no rain penalty</em>
      </div>
      <div class="selection-rule" data-level="light">
        <strong>Light or uncertain</strong><span>Meaningful rain below the likely threshold</span><em>Middle rain-adjusted model</em>
      </div>
      <div class="selection-rule" data-level="likely">
        <strong>Likely rain</strong><span>At least 60% probability and 0.3 mm</span><em>Lower-middle rain-adjusted model</em>
      </div>
      <div class="selection-rule" data-level="heavy">
        <strong>Heavy rain</strong><span>At least 80% probability and 1.0 mm</span><em>Lowest rain-adjusted model</em>
      </div>
      <div class="selection-rule" data-level="observed">
        <strong>Satellite nowcast</strong><span>Completed hour with satellite sunlight data</span><em>Hourly rain penalty is still applied</em>
      </div>
      <div class="selection-rule" data-level="actual">
        <strong>Inverter actual</strong><span>Workbook contains readings for the hour</span><em>Measured kWh replaces forecast for the covered time</em>
      </div>
      <p>Rain conditions use the middle probability and rainfall amount reported by the available weather models for that hour.</p>
    </div>
  </section>

  <section class="forecast-export battery-card" id="batteryForecastCard">
    <div class="export-heading">
      <div>
        <div class="export-kicker">Battery Charging Forecast</div>
        <div class="export-title">{safe_name}</div>
      </div>
      <button type="button" class="card-export-btn" id="batteryDownloadBtn" title="Save battery charging forecast as JPEG">
        <span class="export-download-icon" aria-hidden="true">&#8595;</span>
        Save JPEG
      </button>
    </div>

    <div class="battery-controls">
      <div class="battery-control">
        <label for="batteryDate">Forecast Date</label>
        <select id="batteryDate"></select>
      </div>
      <div class="battery-control">
        <label for="initialSoc">Initial SOC %</label>
        <input id="initialSoc" type="number" min="0" max="100" step="1" value="{min(max(initial_soc_percent, 0.0), 100.0):.0f}" />
      </div>
      <div class="battery-control">
        <label for="consumptionKwh">Approx. kWh Consumption / Hour</label>
        <input id="consumptionKwh" type="number" min="0" step="0.1" value="{max(0.0, consumption_kwh_per_hour):.1f}" />
      </div>
      <label class="battery-check" for="chargeFirst">
        <input id="chargeFirst" type="checkbox"{' checked' if charge_first else ''} />
        Charge first
      </label>
    </div>

    <div class="battery-series-controls" aria-label="Battery chart plots">
      <label class="battery-series-toggle" for="showSolarPlot">
        <input id="showSolarPlot" type="checkbox" checked />
        <span class="battery-series-swatch" style="--series-color:#2563EB" aria-hidden="true"></span>
        Forecasted solar
      </label>
      <label class="battery-series-toggle" for="showConsumptionPlot">
        <input id="showConsumptionPlot" type="checkbox" checked />
        <span class="battery-series-swatch" style="--series-color:#DC2626" aria-hidden="true"></span>
        Approx. consumption
      </label>
    </div>

    <div class="battery-chart-wrap">
      <svg class="battery-chart" id="batterySocChart" viewBox="0 0 720 280" role="img" aria-label="Forecasted battery state of charge, hourly solar production, and approximate consumption from 6am to 6pm"></svg>
    </div>
    <div class="battery-summary">
      <span>Forecasted SOC: <strong id="batteryEndSoc">0%</strong></span>
      <span>Capacity: <strong>{max(0.0, battery_capacity_ah):,.0f} Ah</strong></span>
      <span>Max charge: <strong>{max(0.0, max_charging_ampere):,.0f} A</strong></span>
      <span>Stop discharge: <strong>{min(max(stop_discharge_soc_percent, 0.0), 100.0):.0f}%</strong></span>
      <span>Source: <strong>{mode_name} hourly</strong></span>
    </div>
  </section>

  <section class="forecast-export history-card" id="productionHistoryCard">
    <div class="history-heading">
      <div>
        <div class="export-kicker">Forecast &amp; Nowcast History</div>
        <div class="export-title">{safe_name}</div>
      </div>
      <div class="export-actions">
        <nav class="month-nav" aria-label="Production history month">
          <a href="{html_lib.escape(history_url(previous_month))}" aria-label="Previous month" title="Previous month">&#8249;</a>
          <strong>{history_month.strftime('%B %Y')}</strong>
          <a href="{html_lib.escape(history_url(next_month))}" aria-label="Next month" title="Next month">&#8250;</a>
        </nav>
        <button type="button" class="card-export-btn" id="historyDownloadBtn" title="Save forecast and nowcast history as JPEG">
          <span class="export-download-icon" aria-hidden="true">&#8595;</span>
          Save JPEG
        </button>
      </div>
    </div>
    <div class="calendar-scroll">
      <div class="calendar-grid">
        <div class="weekday">Mon</div><div class="weekday">Tue</div><div class="weekday">Wed</div>
        <div class="weekday">Thu</div><div class="weekday">Fri</div><div class="weekday">Sat</div>
        <div class="weekday">Sun</div>
        {history_cells}
      </div>
    </div>
    <div class="history-summary">{history_summary}</div>
  </section>
</main>
<script src="https://cdn.jsdelivr.net/npm/html2canvas@1.4.1/dist/html2canvas.min.js"></script>
<script>
const siteName = {site_name_json};
const batteryForecast = {battery_forecast_json};
const hourlyForecast = {hourly_forecast_json};
const batteryDate = document.getElementById("batteryDate");
const initialSoc = document.getElementById("initialSoc");
const consumptionKwh = document.getElementById("consumptionKwh");
const chargeFirst = document.getElementById("chargeFirst");
const showSolarPlot = document.getElementById("showSolarPlot");
const showConsumptionPlot = document.getElementById("showConsumptionPlot");
const batteryChart = document.getElementById("batterySocChart");
const batteryEndSoc = document.getElementById("batteryEndSoc");
const hourlyDate = document.getElementById("hourlyDate");
const hourlyRows = document.getElementById("hourlyRows");
const hourlyCard = document.getElementById("hourlyForecastCard");
const dayBars = Array.from(document.querySelectorAll(".day-bar"));

function safeFilePart(value, fallback) {{
  return String(value || "")
    .toLowerCase()
    .replace(/[^a-z0-9]+/g, "-")
    .replace(/(^-|-$)/g, "") || fallback;
}}

function downloadSectionAsJpeg(targetId, filenamePart, trigger = null) {{
  const target = document.getElementById(targetId);
  if (!target || typeof html2canvas !== "function") return;

  const scrollRegions = Array.from(target.querySelectorAll(
    ".hourly-table-wrap, .battery-chart-wrap, .calendar-scroll"
  ));
  const originalStyles = scrollRegions.map(region => ({{
    region,
    overflow: region.style.overflow,
    width: region.style.width
  }}));
  scrollRegions.forEach(region => {{
    const fullWidth = region.scrollWidth;
    region.style.overflow = "visible";
    region.style.width = `${{fullWidth}}px`;
  }});
  target.classList.add("jpeg-exporting");
  if (trigger) {{
    trigger.disabled = true;
    trigger.setAttribute("aria-busy", "true");
  }}

  requestAnimationFrame(() => {{
    const captureWidth = Math.max(target.scrollWidth, target.offsetWidth);
    html2canvas(target, {{
      backgroundColor: "#ffffff",
      scale: 2,
      useCORS: true,
      width: captureWidth,
      windowWidth: captureWidth,
      height: target.scrollHeight,
      windowHeight: target.scrollHeight
    }}).then(canvas => {{
      const link = document.createElement("a");
      const safeName = safeFilePart(siteName, "solar-site");
      link.download = `${{safeName}}-${{filenamePart}}.jpg`;
      link.href = canvas.toDataURL("image/jpeg", 0.95);
      link.click();
    }}).catch(error => {{
      console.error("JPEG export failed", error);
      window.alert("The JPEG could not be saved. Please try again.");
    }}).finally(() => {{
      originalStyles.forEach(({{ region, overflow, width }}) => {{
        region.style.overflow = overflow;
        region.style.width = width;
      }});
      target.classList.remove("jpeg-exporting");
      if (trigger) {{
        trigger.disabled = false;
        trigger.removeAttribute("aria-busy");
      }}
    }});
  }});
}}

document.getElementById("siteDownloadBtn").addEventListener("click", function (e) {{
  e.preventDefault();
  const d = new Date().toISOString().slice(0, 10);
  downloadSectionAsJpeg("solarSiteCapture", `solar-forecast-${{d}}`, this);
}});

document.getElementById("hourlyDownloadBtn").addEventListener("click", function () {{
  const day = hourlyForecast.days?.[Number(hourlyDate.value) || 0];
  const datePart = safeFilePart(day?.date, "selected-date");
  downloadSectionAsJpeg("hourlyForecastCard", `hourly-prediction-${{datePart}}`, this);
}});

document.getElementById("batteryDownloadBtn").addEventListener("click", function () {{
  const day = batteryForecast.days?.[Number(batteryDate.value) || 0];
  const datePart = safeFilePart(day?.date, "selected-date");
  downloadSectionAsJpeg("batteryForecastCard", `battery-charging-${{datePart}}`, this);
}});

document.getElementById("historyDownloadBtn").addEventListener("click", function () {{
  downloadSectionAsJpeg(
    "productionHistoryCard",
    "forecast-nowcast-history-{history_month.strftime('%Y-%m')}",
    this
  );
}});

function renderHourlyForecast() {{
  const selectedIndex = Number(hourlyDate.value) || 0;
  const day = (hourlyForecast.days || [])[selectedIndex];
  hourlyRows.innerHTML = "";
  dayBars.forEach((bar, index) => {{
    bar.setAttribute("aria-pressed", index === selectedIndex ? "true" : "false");
  }});
  if (!day) return;

  (day.rows || []).forEach((row) => {{
    const tr = document.createElement("tr");
    const rainText = `${{Math.round(row.rainProbability)}}% / ${{Number(row.rainMm).toFixed(1)}} mm`;
    const effectText = row.effectLabel || (row.rainPenaltyPercent > 0
        ? `-${{Math.round(row.rainPenaltyPercent)}}%`
        : row.rainLevel === "dry" || row.rainLevel === "off"
          ? "None"
          : "No extra penalty");
    const values = [
      row.time,
      `${{Number(row.expectedKwh).toFixed(1)}} kWh`,
      `${{Number(row.conservativeKwh).toFixed(1)}}-${{Number(row.upperKwh).toFixed(1)}} kWh`,
      rainText,
      effectText,
      row.model,
      row.interpretation
    ];
    values.forEach((value, index) => {{
      const td = document.createElement("td");
      td.textContent = value;
      if (index === 3) {{
        td.className = "rain-signal";
        td.dataset.level = row.rainLevel;
      }} else if (index === 6) {{
        td.className = "hourly-interpretation";
      }}
      tr.appendChild(td);
    }});
    hourlyRows.appendChild(tr);
  }});
}}

function selectHourlyDay(index, shouldScroll = false) {{
  const dayCount = (hourlyForecast.days || []).length;
  if (!dayCount) return;
  const selectedIndex = Math.min(Math.max(Number(index) || 0, 0), dayCount - 1);
  hourlyDate.value = String(selectedIndex);
  renderHourlyForecast();
  if (shouldScroll) {{
    hourlyCard.scrollIntoView({{ behavior: "smooth", block: "start" }});
  }}
}}

function initHourlyDates() {{
  hourlyDate.innerHTML = "";
  (hourlyForecast.days || []).forEach((day, index) => {{
    const option = document.createElement("option");
    option.value = String(index);
    option.textContent = day.label;
    hourlyDate.appendChild(option);
  }});
  selectHourlyDay(0);
}}

function clampNumber(value, min, max) {{
  const n = Number(value);
  if (!Number.isFinite(n)) return min;
  return Math.min(Math.max(n, min), max);
}}

function svgEl(name, attrs = {{}}, text = "") {{
  const el = document.createElementNS("http://www.w3.org/2000/svg", name);
  Object.entries(attrs).forEach(([key, value]) => el.setAttribute(key, value));
  if (text) el.textContent = text;
  return el;
}}

function initBatteryDates() {{
  batteryDate.innerHTML = "";
  (batteryForecast.days || []).forEach((day, index) => {{
    const option = document.createElement("option");
    option.value = String(index);
    option.textContent = day.label;
    batteryDate.appendChild(option);
  }});
}}

function estimateSoc(day) {{
  const capacity = Math.max(Number(batteryForecast.capacityAh) || 0, 0)
    * Math.max(Number(batteryForecast.nominalVoltage) || 0, 0) / 1000;
  const initialPct = clampNumber(initialSoc.value, 0, 100);
  const consumptionPerHour = Math.max(Number(consumptionKwh.value) || 0, 0);
  const shouldChargeFirst = chargeFirst.checked;
  const points = day?.points || [];
  const maxChargeKwh = Math.max(Number(batteryForecast.maxChargingAmpere) || 0, 0)
    * Math.max(Number(batteryForecast.nominalVoltage) || 0, 0) / 1000;
  const stopDischargePct = clampNumber(batteryForecast.stopDischargeSocPercent, 0, 100);
  const reserveKwh = capacity * (stopDischargePct / 100);
  let socKwh = capacity * (initialPct / 100);

  return points.map((point, index) => {{
    if (index === 0) {{
      const socPct = capacity > 0 ? (socKwh / capacity) * 100 : 0;
      return {{
        ...point,
        socPct,
        chargeKwh: 0,
        unmetLoadKwh: 0,
        loadKwh: 0,
        netSolarKwh: 0,
        isInitial: true
      }};
    }}

    const interval = point;
    const production = Math.max(Number(interval.productionKwh) || 0, 0);
    const availableCapacityKwh = Math.max(capacity - socKwh, 0);
    const netSolarKwh = production - consumptionPerHour;
    let solarForLoadKwh = shouldChargeFirst ? 0 : Math.min(production, consumptionPerHour);
    const chargeableSolarKwh = shouldChargeFirst
      ? production
      : Math.max(production - solarForLoadKwh, 0);
    const chargeKwh = chargeableSolarKwh > 0 && maxChargeKwh > 0
      ? Math.min(chargeableSolarKwh, maxChargeKwh, availableCapacityKwh)
      : 0;
    if (shouldChargeFirst) {{
      solarForLoadKwh = Math.min(Math.max(production - chargeKwh, 0), consumptionPerHour);
    }}
    const loadShortfallKwh = Math.max(consumptionPerHour - solarForLoadKwh, 0);
    const dischargeableKwh = Math.max(socKwh - reserveKwh, 0);
    const blockBatteryDischarge = shouldChargeFirst && chargeKwh > 0;
    const dischargeKwh = blockBatteryDischarge
      ? 0
      : Math.min(loadShortfallKwh, dischargeableKwh);
    const unmetLoadKwh = Math.max(loadShortfallKwh - dischargeKwh, 0);
    socKwh = Math.min(Math.max(socKwh + chargeKwh - dischargeKwh, 0), capacity || 0);
    const socPct = capacity > 0 ? (socKwh / capacity) * 100 : 0;
    return {{ ...point, socPct, chargeKwh, dischargeKwh, unmetLoadKwh, loadKwh: consumptionPerHour, netSolarKwh }};
  }});
}}

function renderBatteryChart() {{
  const day = batteryForecast.days?.[Number(batteryDate.value) || 0];
  const points = estimateSoc(day);
  batteryChart.innerHTML = "";

  const width = 720;
  const height = 280;
  const left = 48;
  const right = 52;
  const top = 34;
  const bottom = 42;
  const plotW = width - left - right;
  const plotH = height - top - bottom;

  [0, 25, 50, 75, 100].forEach(tick => {{
    const y = top + plotH - (tick / 100) * plotH;
    batteryChart.appendChild(svgEl("line", {{
      x1: left, y1: y, x2: width - right, y2: y,
      stroke: "#CBD5E1", "stroke-width": "1", "stroke-dasharray": "3 5"
    }}));
    batteryChart.appendChild(svgEl("text", {{
      x: left - 8, y: y + 4, "text-anchor": "end",
      "font-size": "12", "font-weight": "800", fill: "#475569"
    }}, `${{tick}}%`));
  }});

  const stopDischargePct = clampNumber(batteryForecast.stopDischargeSocPercent, 0, 100);
  const reserveY = top + plotH - (stopDischargePct / 100) * plotH;
  batteryChart.appendChild(svgEl("line", {{
    x1: left, y1: reserveY, x2: width - right, y2: reserveY,
    stroke: "#DC2626", "stroke-width": "2", "stroke-dasharray": "7 5"
  }}));
  batteryChart.appendChild(svgEl("text", {{
    x: width - right, y: Math.max(top + 11, reserveY - 5), "text-anchor": "end",
    "font-size": "11", "font-weight": "900", fill: "#B91C1C"
  }}, `Stop ${{Math.round(stopDischargePct)}}%`));

  batteryChart.appendChild(svgEl("text", {{
    x: left - 8, y: 15, "text-anchor": "end",
    "font-size": "12", "font-weight": "900", fill: "#334155"
  }}, "SOC"));

  if (!points.length) {{
    batteryChart.appendChild(svgEl("text", {{
      x: width / 2, y: height / 2, "text-anchor": "middle",
      "font-size": "14", "font-weight": "900", fill: "#64748B"
    }}, "No battery forecast data"));
    batteryEndSoc.textContent = "0%";
    return;
  }}

  const coords = points.map((point, index) => {{
    const x = left + (points.length === 1 ? 0 : (index / (points.length - 1)) * plotW);
    const y = top + plotH - (clampNumber(point.socPct, 0, 100) / 100) * plotH;
    return {{ x, y, point, index }};
  }});

  const hourlyPoints = points.filter(point => !point.isInitial);
  const energyValues = hourlyPoints.flatMap(point => [
    Math.max(Number(point.productionKwh) || 0, 0),
    Math.max(Number(point.loadKwh) || 0, 0)
  ]);
  const largestEnergyValue = Math.max(...energyValues, 1);
  const energyMagnitude = 10 ** Math.floor(Math.log10(largestEnergyValue));
  const normalizedEnergy = largestEnergyValue / energyMagnitude;
  const niceEnergyFactor = normalizedEnergy <= 1
    ? 1
    : normalizedEnergy <= 2
      ? 2
      : normalizedEnergy <= 5
        ? 5
        : 10;
  const energyMax = niceEnergyFactor * energyMagnitude;
  const showEnergyAxis = showSolarPlot.checked || showConsumptionPlot.checked;

  if (showEnergyAxis) {{
    [0, 25, 50, 75, 100].forEach(tick => {{
      const y = top + plotH - (tick / 100) * plotH;
      const value = (tick / 100) * energyMax;
      batteryChart.appendChild(svgEl("text", {{
        x: width - right + 8, y: y + 4, "text-anchor": "start",
        "font-size": "11", "font-weight": "800", fill: "#475569"
      }}, Number.isInteger(value) ? value.toFixed(0) : value.toFixed(1)));
    }});
    batteryChart.appendChild(svgEl("text", {{
      x: width - right + 8, y: 15, "text-anchor": "start",
      "font-size": "12", "font-weight": "900", fill: "#334155"
    }}, "kWh"));
  }}

  function drawHourlyEnergySeries(valueKey, color) {{
    if (!hourlyPoints.length) return;
    const intervals = hourlyPoints.map((point, index) => {{
      const value = Math.max(Number(point[valueKey]) || 0, 0);
      return {{
        point,
        value,
        startX: left + (index / hourlyPoints.length) * plotW,
        endX: left + ((index + 1) / hourlyPoints.length) * plotW,
        y: top + plotH - (value / energyMax) * plotH
      }};
    }});
    const pathParts = [`M ${{intervals[0].startX.toFixed(2)}} ${{intervals[0].y.toFixed(2)}}`];
    intervals.forEach((interval, index) => {{
      pathParts.push(`L ${{interval.endX.toFixed(2)}} ${{interval.y.toFixed(2)}}`);
      if (index < intervals.length - 1) {{
        pathParts.push(`L ${{interval.endX.toFixed(2)}} ${{intervals[index + 1].y.toFixed(2)}}`);
      }}
    }});
    batteryChart.appendChild(svgEl("path", {{
      d: pathParts.join(" "), fill: "none", stroke: color, "stroke-width": "3",
      "stroke-linecap": "round", "stroke-linejoin": "round"
    }}));
  }}

  if (showSolarPlot.checked) {{
    drawHourlyEnergySeries("productionKwh", "#2563EB");
  }}
  if (showConsumptionPlot.checked) {{
    drawHourlyEnergySeries("loadKwh", "#DC2626");
  }}

  batteryChart.appendChild(svgEl("polyline", {{
    points: coords.map(p => `${{p.x.toFixed(2)}},${{p.y.toFixed(2)}}`).join(" "),
    fill: "none", stroke: "#F97316", "stroke-width": "4",
    "stroke-linecap": "round", "stroke-linejoin": "round"
  }}));

  coords.forEach(({{ x, y, point, index }}) => {{
    const marker = svgEl("circle", {{
      cx: x, cy: y, r: "4", fill: "#FBBF24", stroke: "#92400E", "stroke-width": "1"
    }});
    marker.appendChild(svgEl(
      "title",
      {{}},
      point.isInitial
        ? `${{point.time}}: initial ${{Math.round(point.socPct)}}% SOC`
        : `${{point.time}}: ${{Math.round(point.socPct)}}% SOC | solar ${{point.productionKwh.toFixed(2)}} kWh | load ${{point.loadKwh.toFixed(2)}} kWh | charge ${{point.chargeKwh.toFixed(2)}} kWh | discharge ${{point.dischargeKwh.toFixed(2)}} kWh | unmet load ${{point.unmetLoadKwh.toFixed(2)}} kWh`
    ));
    batteryChart.appendChild(marker);
    batteryChart.appendChild(svgEl("text", {{
      x, y: Math.max(top + 10, y - 10), "text-anchor": "middle",
      "font-size": "12", "font-weight": "900", fill: "#111827"
    }}, `${{Math.round(point.socPct)}}%`));
    if (index % 2 === 0 || index === coords.length - 1) {{
      batteryChart.appendChild(svgEl("text", {{
        x, y: height - 14, "text-anchor": "middle",
        "font-size": "12", "font-weight": "800", fill: "#475569"
      }}, point.time));
    }}
  }});

  const last = coords[coords.length - 1].point.socPct;
  batteryEndSoc.textContent = `${{Math.round(last)}}%`;
}}

initBatteryDates();
initHourlyDates();
[batteryDate, initialSoc, consumptionKwh, chargeFirst, showSolarPlot, showConsumptionPlot].forEach(el => el.addEventListener("input", renderBatteryChart));
[batteryDate, initialSoc, consumptionKwh, chargeFirst, showSolarPlot, showConsumptionPlot].forEach(el => el.addEventListener("change", renderBatteryChart));
hourlyDate.addEventListener("change", () => selectHourlyDay(hourlyDate.value));
dayBars.forEach((bar, index) => {{
  bar.addEventListener("click", () => selectHourlyDay(index, true));
}});
renderBatteryChart();
</script>
</body>
</html>
"""



# ---------------- FLASK APP ----------------

app = Flask(__name__)
cache_init()

@app.get("/")
def index():
    # Query params with sane defaults
    tz = request.args.get("tz", DEFAULT_TZ)
    country = request.args.get("country", DEFAULT_COUNTRY)  # kept for URL/cache compatibility
    model = request.args.get("model", DEFAULT_MODEL)
    list_id = request.args.get("list")
    view = normalized_view(request.args.get("view"))

    # Cache bypass
    nocache = request.args.get("nocache") == "1"

    # Keep cache tidy
    cache_prune()

    # "Today" is based on requested timezone
    qdate = safe_now_date_in_tz(tz)
    min_date = qdate
    max_date = qdate + timedelta(days=FUTURE_DAYS_ALLOWED)

    # Target date (date being queried)
    date_str = request.args.get("date")
    if date_str:
        try:
            target_date = datetime.strptime(date_str, "%Y-%m-%d").date()
        except ValueError:
            return Response("Invalid date format. Use YYYY-MM-DD.", status=400, mimetype="text/plain")

        if not (min_date <= target_date <= max_date):
            return Response(
                f"Date out of allowed range. Use {min_date.isoformat()} to {max_date.isoformat()} (tz={tz}).",
                status=400,
                mimetype="text/plain",
            )
    else:
        target_date = qdate

    # Places file (NEW: coordinates-based)
    try:
        places, lists_info, active_list_id = read_places_file(DEFAULT_PLACES_FILE, list_id)
        p_sig = f"{places_signature(DEFAULT_PLACES_FILE, active_list_id)}:{view}:{SOLAR_TOTAL_CACHE_VERSION}"
    except FileNotFoundError:
        return Response(
            f"Missing places file: {DEFAULT_PLACES_FILE}\n"
            "Create it as a valid JSON.\n",
            status=500,
            mimetype="text/plain",
        )
    except Exception as e:
        return Response(f"Error reading places file: {e}", status=500, mimetype="text/plain")

    if not places:
        return Response(
            f"No places found in {DEFAULT_PLACES_FILE}. Add lines like:\nAIVR, 13.174, 121.278\n",
            status=500,
            mimetype="text/plain",
        )

    # Base params for nav buttons
    base_params = {
        "tz": tz,
        "country": country,
        "model": model,
        "list": active_list_id
    }
    if view == "solar":
        base_params["view"] = "solar"
    if nocache:
        base_params["nocache"] = "1"

    # Try cache first (unless nocache)
    if not nocache:
        cached_html = cache_get(
            query_date=qdate.isoformat(),
            target_date=target_date.isoformat(),
            tz=tz,
            country=country,
            model=model,
            places_sig=p_sig,
        )
        if cached_html is not None:
            return Response(cached_html, mimetype="text/html")

    # ---------------- LIVE BUILD ----------------

    client = OpenMeteoClient()

    # cell_map[place_label][hour_key] = (icon, precip_mm, pop_pct, solar_wm2)
    cell_map: Dict[str, Dict[str, Tuple[str, float, int, float]]] = {}
    time_index: List[datetime] = []
    seen_hours = set()

    for p in places:
        if view == "solar":
            hourly = client.hybrid_solar_forecast(p.lat, p.lon, tz, model)
        else:
            hourly = client.hourly_forecast(p.lat, p.lon, tz, model)
        cells: Dict[str, Tuple[str, float, int, float]] = {}

        for t, pr, pop, cc, solar in zip(
            hourly["time"],
            hourly["precip"],
            hourly["pop"],
            hourly["cloud"],
            hourly["solar"],
        ):
            if t.date() != target_date:
                continue

            hk = t.strftime("%H:00")
            if hk not in seen_hours:
                seen_hours.add(hk)
                time_index.append(t)

            cells[hk] = (weather_icon(cc, pr, t), pr, int(pop or 0), _to_float(solar))

        cell_map[p.label] = cells

    # Sort the hours left-to-right
    time_index.sort(key=lambda x: x)

    html = render_html(
        target_date=target_date,
        min_date=min_date,
        max_date=max_date,
        tz=tz,
        model=model,
        places=places,
        lists_info=lists_info,
        active_list_id=active_list_id,
        time_index=time_index,
        cell_map=cell_map,
        from_cache=False,
        nocache=nocache,
        base_params=base_params,
        view=view,
    )

    # Store in cache (only if not nocache)
    if not nocache:
        cache_put(
            query_date=qdate.isoformat(),
            target_date=target_date.isoformat(),
            tz=tz,
            country=country,
            model=model,
            places_sig=p_sig,
            html=html,
        )

    return Response(html, mimetype="text/html")


@app.get("/solar-site")
def solar_site_page():
    tz = request.args.get("tz", DEFAULT_TZ)
    model = request.args.get("model", DEFAULT_MODEL)
    selected_site_id = (request.args.get("site") or "").strip()
    rain_values = request.args.getlist("rain")
    rain_effect = "1" in rain_values if rain_values else True
    today = safe_now_date_in_tz(tz)

    history_month_raw = (request.args.get("history_month") or "").strip()
    try:
        history_month = datetime.strptime(history_month_raw, "%Y-%m").date().replace(day=1)
        if not (2000 <= history_month.year <= today.year + 1):
            history_month = today.replace(day=1)
    except ValueError:
        history_month = today.replace(day=1)

    try:
        sites = read_sites_file(DEFAULT_PLACES_FILE)
    except Exception:
        sites = []

    selected_site = next(
        (site for site in sites if str(site.get("id", "")).strip() == selected_site_id),
        None,
    )
    if not selected_site_id and sites:
        selected_site = sites[0]
        selected_site_id = str(selected_site.get("id", "")).strip()

    default_name = str(selected_site.get("name", "Solar Site")) if selected_site else "Solar Site"
    default_lat = _to_float(selected_site.get("lat") if selected_site else None, 13.174000)
    default_lon = _to_float(selected_site.get("lon") if selected_site else None, 121.278000)
    default_kw = _to_float(
        selected_site.get("system_kw", selected_site.get("system_size_kw", selected_site.get("kw"))) if selected_site else None,
        95.84,
    )
    default_eff = _to_float(
        selected_site.get("efficiency_factor", selected_site.get("efficiency", selected_site.get("eff"))) if selected_site else None,
        0.80,
    )
    default_battery_capacity_ah = _to_float(
        selected_site.get(
            "total_battery_capacity_ah",
            selected_site.get("battery_capacity_ah", selected_site.get("total_battery_capacity_kwh", selected_site.get("battery_capacity_kwh"))),
        ) if selected_site else None,
        100.0,
    )
    default_max_charging_ampere = _to_float(
        selected_site.get("max_charging_ampere", selected_site.get("max_charge_ampere")) if selected_site else None,
        100.0,
    )
    default_initial_soc_percent = _to_float(
        selected_site.get("initial_soc_percent", selected_site.get("initial_soc")) if selected_site else None,
        20.0,
    )
    default_stop_discharge_soc_percent = _to_float(
        selected_site.get("stop_discharge_soc_percent", selected_site.get("stop_discharge_soc")) if selected_site else None,
        20.0,
    )
    default_consumption_kwh_per_hour = _to_float(
        selected_site.get("consumption_kwh_per_hour", selected_site.get("hourly_consumption_kwh")) if selected_site else None,
        0.0,
    )
    default_charge_first = bool(selected_site.get("charge_first", False)) if selected_site else False

    site_name = default_name.strip() or "Solar Site"
    lat = default_lat
    lon = default_lon
    system_kw = max(0.0, default_kw)
    efficiency_factor = normalize_efficiency_factor(default_eff)
    battery_capacity_ah = max(0.0, default_battery_capacity_ah)
    max_charging_ampere = max(0.0, default_max_charging_ampere)
    initial_soc_percent = min(max(default_initial_soc_percent, 0.0), 100.0)
    stop_discharge_soc_percent = min(max(default_stop_discharge_soc_percent, 0.0), 100.0)
    consumption_kwh_per_hour = max(0.0, default_consumption_kwh_per_hour)
    charge_first = default_charge_first
    inverter_data_file = str(selected_site.get("inverter_data_file") or "").strip() if selected_site else ""
    if inverter_data_file and not os.path.isabs(inverter_data_file):
        inverter_data_file = os.path.join(
            os.path.dirname(os.path.abspath(DEFAULT_PLACES_FILE)),
            inverter_data_file,
        )
    try:
        inverter_data = inverter_actuals_for_file(inverter_data_file)
    except Exception:
        inverter_data = {"hourly": {}, "daily": {}, "last_reading": None, "source": ""}
    inverter_hourly = inverter_data.get("hourly") or {}
    inverter_daily = inverter_data.get("daily") or {}
    inverter_last_reading = inverter_data.get("last_reading")
    inverter_note = ""
    if isinstance(inverter_last_reading, datetime):
        inverter_note = (
            f"Inverter feed: {inverter_data.get('source') or 'workbook'} through "
            f"{inverter_last_reading.strftime('%b %d, %I:%M %p').replace(' 0', ' ')}"
        )

    if selected_site is None and selected_site_id:
        selected_site_id = ""

    if not (-90.0 <= lat <= 90.0):
        return Response("Invalid latitude. Use a value from -90 to 90.", status=400, mimetype="text/plain")
    if not (-180.0 <= lon <= 180.0):
        return Response("Invalid longitude. Use a value from -180 to 180.", status=400, mimetype="text/plain")

    try:
        weather_client = OpenMeteoClient()
        hourly = weather_client.hybrid_solar_forecast(lat, lon, tz, model)
        hourly = weather_client.add_rain_model_scenarios(hourly, lat, lon, tz, model)
    except Exception as e:
        return Response(f"Error fetching forecast: {e}", status=502, mimetype="text/plain")

    base_calibration_factor, base_calibration_days = solar_calibration_factor(
        selected_site_id,
        False,
        today,
    )
    rain_calibration_factor, rain_calibration_days = solar_calibration_factor(
        selected_site_id,
        True,
        today,
    )
    calibration_factor = rain_calibration_factor if rain_effect else base_calibration_factor
    calibration_days = rain_calibration_days if rain_effect else base_calibration_days

    by_day: Dict[date, Dict[str, float]] = {}
    hourly_production: List[Tuple[datetime, float, float, float]] = []
    hourly_detail_rows: List[Dict[str, object]] = []
    solar_sources = hourly.get("solar_source") or ["forecast"] * len(hourly["time"])
    rain_solar_values = hourly.get("rain_solar") or hourly["solar"]
    conservative_solar_values = hourly.get("rain_conservative_solar") or rain_solar_values
    upper_solar_values = hourly.get("rain_upper_solar") or rain_solar_values
    consensus_values = hourly.get("rain_consensus_used") or [False] * len(hourly["time"])
    adjustment_precip_values = hourly.get("rain_adjustment_precip") or hourly["precip"]
    adjustment_pop_values = hourly.get("rain_adjustment_pop") or hourly["pop"]
    selected_model_values = hourly.get("rain_selected_model") or ["Primary forecast"] * len(hourly["time"])
    interpretation_values = hourly.get("rain_interpretation") or ["Primary solar forecast used."] * len(hourly["time"])
    rain_level_values = hourly.get("rain_level") or ["dry"] * len(hourly["time"])
    retained_factor_values = hourly.get("rain_retained_factor") or [1.0] * len(hourly["time"])
    rain_model_count = max(1, int(_to_float(hourly.get("rain_model_count"), 1.0)))
    completed_hour = safe_now_in_tz(tz).replace(minute=0, second=0, microsecond=0)
    has_today_inverter_feed = today in inverter_daily
    for t, solar, rain_solar, conservative_solar, upper_solar, precipitation, probability, adjustment_precipitation, adjustment_probability, selected_model, interpretation, rain_level, retained_factor, solar_source, consensus_used in zip(
        hourly["time"],
        hourly["solar"],
        rain_solar_values,
        conservative_solar_values,
        upper_solar_values,
        hourly["precip"],
        hourly["pop"],
        adjustment_precip_values,
        adjustment_pop_values,
        selected_model_values,
        interpretation_values,
        rain_level_values,
        retained_factor_values,
        solar_sources,
        consensus_values,
    ):
        base_kwh = hourly_solar_production_kwh(
            system_kw,
            _to_float(solar),
            efficiency_factor,
        )
        rain_kwh = hourly_solar_production_kwh(
            system_kw,
            _to_float(rain_solar),
            efficiency_factor,
        )
        conservative_kwh = hourly_solar_production_kwh(
            system_kw,
            _to_float(conservative_solar),
            efficiency_factor,
        )
        upper_kwh = hourly_solar_production_kwh(
            system_kw,
            _to_float(upper_solar),
            efficiency_factor,
        )
        hour_key = t.replace(minute=0, second=0, microsecond=0)
        actual_record = inverter_hourly.get(hour_key) if t <= completed_hour else None
        actual_kwh = max(0.0, _to_float(actual_record.get("kwh"))) if actual_record else 0.0
        actual_fraction = (
            min(max(_to_float(actual_record.get("coverage_minutes")) / 60.0, 0.0), 1.0)
            if actual_record
            else 0.0
        )
        forecast_fraction = 1.0 - actual_fraction
        base_hybrid_display_kwh = actual_kwh + (
            base_kwh * base_calibration_factor * forecast_fraction
        )
        rain_hybrid_display_kwh = actual_kwh + (
            rain_kwh * rain_calibration_factor * forecast_fraction
        )
        conservative_hybrid_display_kwh = actual_kwh + (
            conservative_kwh * rain_calibration_factor * forecast_fraction
        )
        upper_hybrid_display_kwh = actual_kwh + (
            upper_kwh * rain_calibration_factor * forecast_fraction
        )
        day_totals = by_day.setdefault(
            t.date(),
            {
                "irradiation_wh_m2": 0.0,
                "base_forecast_kwh": 0.0,
                "rain_forecast_kwh": 0.0,
                "rain_conservative_kwh": 0.0,
                "rain_upper_kwh": 0.0,
                "precipitation_mm": 0.0,
                "rain_hours": 0.0,
                "has_nowcast": 0.0,
                "consensus_hours": 0.0,
                "estimated_by_now_kwh": 0.0,
                "base_hybrid_display_kwh": 0.0,
                "rain_hybrid_display_kwh": 0.0,
                "conservative_hybrid_display_kwh": 0.0,
                "upper_hybrid_display_kwh": 0.0,
                "inverter_actual_kwh": 0.0,
                "has_inverter_actual": 0.0,
                "rain_model_count": float(rain_model_count),
            },
        )
        day_totals["irradiation_wh_m2"] += max(0.0, _to_float(solar))
        day_totals["base_forecast_kwh"] += base_kwh
        day_totals["rain_forecast_kwh"] += rain_kwh
        day_totals["rain_conservative_kwh"] += conservative_kwh
        day_totals["rain_upper_kwh"] += upper_kwh
        day_totals["base_hybrid_display_kwh"] += base_hybrid_display_kwh
        day_totals["rain_hybrid_display_kwh"] += rain_hybrid_display_kwh
        day_totals["conservative_hybrid_display_kwh"] += conservative_hybrid_display_kwh
        day_totals["upper_hybrid_display_kwh"] += upper_hybrid_display_kwh
        day_totals["inverter_actual_kwh"] += actual_kwh
        if actual_fraction > 0:
            day_totals["has_inverter_actual"] = 1.0
        day_totals["precipitation_mm"] += max(0.0, _to_float(precipitation))
        if _to_float(precipitation) > 0 or _to_float(probability) >= 50:
            day_totals["rain_hours"] += 1
        if solar_source == "nowcast":
            day_totals["has_nowcast"] = 1.0
        if consensus_used:
            day_totals["consensus_hours"] += 1

        selected_display_kwh = (
            rain_hybrid_display_kwh if rain_effect else base_hybrid_display_kwh
        )
        if t.date() == today and t <= completed_hour:
            day_totals["estimated_by_now_kwh"] += (
                actual_kwh if has_today_inverter_feed else selected_display_kwh
            )
        hourly_production.append((t, selected_display_kwh, base_kwh, rain_kwh))
        actual_source = actual_fraction > 0
        actual_full_hour = actual_fraction >= 0.98
        if actual_source:
            detail_model = (
                "Inverter actual"
                if actual_full_hour
                else f"Inverter + {selected_model if rain_effect else 'primary forecast'}"
            )
            detail_interpretation = (
                "Measured inverter production from the workbook."
                if actual_full_hour
                else f"{actual_fraction * 100.0:.0f}% measured inverter energy; forecast fills the remaining time."
            )
            detail_level = "actual" if actual_full_hour else str(rain_level)
            detail_effect_label = "Measured" if actual_full_hour else f"{actual_fraction * 100.0:.0f}% measured"
        else:
            detail_model = str(selected_model) if rain_effect else "Primary forecast"
            detail_interpretation = (
                str(interpretation)
                if rain_effect
                else "Rain effect is off; the primary solar forecast is used."
            )
            detail_level = str(rain_level) if rain_effect else "off"
            detail_effect_label = ""
        if rain_effect:
            hourly_detail_rows.append(
                {
                    "date": t.date(),
                    "hour": t.hour,
                    "time": t.strftime("%I:%M %p").lstrip("0").lower(),
                    "expectedKwh": round(rain_hybrid_display_kwh, 3),
                    "conservativeKwh": round(conservative_hybrid_display_kwh, 3),
                    "upperKwh": round(upper_hybrid_display_kwh, 3),
                    "rainMm": round(max(0.0, _to_float(adjustment_precipitation)), 2),
                    "rainProbability": round(min(max(_to_float(adjustment_probability), 0.0), 100.0), 1),
                    "rainPenaltyPercent": round((1.0 - min(max(_to_float(retained_factor), 0.0), 1.0)) * 100.0, 1),
                    "rainLevel": detail_level,
                    "model": detail_model,
                    "interpretation": detail_interpretation,
                    "effectLabel": detail_effect_label,
                }
            )
        else:
            hourly_detail_rows.append(
                {
                    "date": t.date(),
                    "hour": t.hour,
                    "time": t.strftime("%I:%M %p").lstrip("0").lower(),
                    "expectedKwh": round(base_hybrid_display_kwh, 3),
                    "conservativeKwh": round(base_hybrid_display_kwh, 3),
                    "upperKwh": round(base_hybrid_display_kwh, 3),
                    "rainMm": round(max(0.0, _to_float(adjustment_precipitation)), 2),
                    "rainProbability": round(min(max(_to_float(adjustment_probability), 0.0), 100.0), 1),
                    "rainPenaltyPercent": 0.0,
                    "rainLevel": detail_level,
                    "model": detail_model,
                    "interpretation": detail_interpretation,
                    "effectLabel": detail_effect_label,
                }
            )

    daily_rows: List[Dict[str, object]] = []
    for day in sorted(by_day.keys())[:7]:
        totals = by_day[day]
        selected_display_kwh = (
            totals["rain_hybrid_display_kwh"]
            if rain_effect
            else totals["base_hybrid_display_kwh"]
        )
        actual_day = inverter_daily.get(day) or {}
        actual_complete = bool(actual_day.get("complete"))
        if actual_complete:
            source_type = "actual"
        elif totals["has_inverter_actual"]:
            source_type = "actual + forecast"
        elif totals["has_nowcast"]:
            source_type = "nowcast"
        else:
            source_type = "forecast"
        daily_rows.append(
            {
                "date": day,
                **totals,
                "base_display_kwh": totals["base_hybrid_display_kwh"],
                "rain_display_kwh": totals["rain_hybrid_display_kwh"],
                "conservative_display_kwh": totals["conservative_hybrid_display_kwh"],
                "upper_display_kwh": totals["upper_hybrid_display_kwh"],
                "production_kwh": selected_display_kwh,
                "remaining_today_kwh": max(
                    selected_display_kwh - totals["estimated_by_now_kwh"],
                    0.0,
                ),
                "source_type": source_type,
                "inverter_actual_complete": actual_complete,
            }
        )

    solar_history_record_forecasts(selected_site_id, daily_rows)
    history_rows = solar_history_for_month(selected_site_id, history_month) if selected_site_id else {}

    hourly_days = []
    for day in sorted(by_day.keys())[:7]:
        hourly_days.append(
            {
                "date": day.isoformat(),
                "label": f"{day.strftime('%b%d')} {day.strftime('%a')}",
                "rows": [
                    row
                    for row in hourly_detail_rows
                    if row["date"] == day and 6 <= int(row["hour"]) <= 18
                ],
            }
        )
    for day in hourly_days:
        for row in day["rows"]:
            row.pop("date", None)
            row.pop("hour", None)

    battery_days: List[Dict[str, object]] = []
    for day in sorted(by_day.keys())[:7]:
        day_start = datetime.combine(day, datetime.min.time()).replace(hour=6)
        points: List[Dict[str, object]] = [
            {
                "time": day_start.strftime("%I%p").lstrip("0").lower(),
                "hour": 6,
                "productionKwh": 0.0,
                "isInitial": True,
            }
        ]
        for t, production_kwh, _base_kwh, _rain_kwh in hourly_production:
            if t.date() != day or not (6 <= t.hour < 18):
                continue
            interval_end = t + timedelta(hours=1)
            points.append(
                {
                    "time": interval_end.strftime("%I%p").lstrip("0").lower(),
                    "hour": interval_end.hour,
                    "forecastHour": t.strftime("%I:%M %p").lstrip("0").lower(),
                    "productionKwh": round(production_kwh, 3),
                    "isInitial": False,
                }
            )
        battery_days.append(
            {
                "date": day.isoformat(),
                "label": f"{day.strftime('%b%d')} {day.strftime('%a')}",
                "points": points,
            }
        )

    html = render_solar_site_html(
        site_name=site_name,
        site_id=selected_site_id,
        sites=sites,
        tz=tz,
        model=model,
        daily_rows=daily_rows,
        battery_days=battery_days,
        hourly_days=hourly_days,
        battery_capacity_ah=battery_capacity_ah,
        max_charging_ampere=max_charging_ampere,
        initial_soc_percent=initial_soc_percent,
        stop_discharge_soc_percent=stop_discharge_soc_percent,
        consumption_kwh_per_hour=consumption_kwh_per_hour,
        charge_first=charge_first,
        rain_effect=rain_effect,
        calibration_factor=calibration_factor,
        calibration_days=calibration_days,
        history_month=history_month,
        history_rows=history_rows,
        inverter_note=inverter_note,
        today=today,
    )
    return Response(html, mimetype="text/html")


@app.get("/admin")
def admin_page():
    # Simple HTML page with Vue.js to manage places
    html = """<!doctype html>
<html>
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1" />
<title>Mindoro Rain Forecast - Admin</title>
<script src="https://cdn.jsdelivr.net/npm/vue@3/dist/vue.global.js"></script>
<style>
body { font-family: Segoe UI, Arial, sans-serif; margin:16px; background:#f5f5f5;}
.container { max-width: 960px; margin: auto; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
h1, h2, h3 { margin-top: 0; }
.form-group { margin-bottom: 10px; }
input { padding: 8px; border: 1px solid #ccc; border-radius: 4px; }
button { padding: 8px 12px; border: none; background: #111; color: white; border-radius: 4px; cursor: pointer; }
button:hover { background: #333; }
button.danger { background: #dc3545; }
button.danger:hover { background: #c82333; }
.list-item, .place-item, .site-item { border: 1px solid #ddd; padding: 10px; margin-bottom: 10px; border-radius: 4px; background: #fafafa; display: flex; align-items: center; gap: 10px;}
.place-item input { flex: 1; }
.site-item { flex-wrap: wrap; }
.site-field { flex: 1 1 120px; display: flex; flex-direction: column; gap: 4px; }
.site-field label { font-size: 11px; font-weight: bold; color: #555; }
.site-field input { width: 100%; box-sizing: border-box; }
.site-field.name-field { flex-basis: 180px; }
.site-field.id-field { flex-basis: 140px; }
.site-check { flex: 0 0 110px; display: flex; align-items: center; gap: 7px; font-size: 12px; font-weight: bold; color: #555; }
.site-check input { width: 16px; height: 16px; }
.actions { margin-left: auto; display: flex; gap: 5px; }
.dragger { cursor: grab; padding: 0 10px; color: #888; }
.section { margin-bottom: 30px; border-bottom: 1px solid #eee; padding-bottom: 20px;}
.section:last-child { border-bottom: none; }
</style>
</head>
<body>
<div id="app" class="container">
  <h1>Admin Page</h1>
  <div v-if="loading">Loading...</div>
  <div v-else>
    <div class="section">
      <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom: 15px;">
        <h2 style="margin:0;">Lists</h2>
        <button @click="addList">Add List</button>
      </div>
      <div class="form-group">
        <label>Default List ID:</label>
        <select v-model="config.default_list" style="padding: 8px;">
            <option v-for="l in config.lists" :value="l.id">{{ l.name }} ({{ l.id }})</option>
        </select>
      </div>
      <div style="display:flex; gap: 10px; flex-wrap:wrap;">
         <div v-for="(l, index) in config.lists" :key="l.id" style="border: 1px solid #ccc; padding: 10px; border-radius: 5px; cursor:pointer;" :style="{background: activeListId === l.id ? '#e0f7fa' : '#fff'}" @click="activeListId = l.id">
            <div style="font-weight:bold;">{{ l.name }}</div>
            <div style="font-size: 12px; color: #666;">ID: {{ l.id }}</div>
            <button class="danger" style="padding: 4px 8px; font-size: 12px; margin-top:5px;" @click.stop="removeList(index)">Delete</button>
         </div>
      </div>
    </div>

    <div class="section" v-if="activeList">
      <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom: 15px;">
          <h2 style="margin:0;">Edit List: <input v-model="activeList.name" placeholder="List Name" /></h2>
      </div>
      <div class="form-group">
        <label style="font-size: 12px; color: #666;">List ID (must be unique):</label>
        <input v-model="activeList.id" placeholder="List ID" style="display:block; margin-top:4px;" />
      </div>

      <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom: 10px;">
          <h3 style="margin:0;">Places</h3>
          <button @click="addPlace">Add Place</button>
      </div>

      <div v-for="(place, index) in activeList.places" :key="index" class="place-item">
         <span class="dragger" @click="movePlaceUp(index)" title="Move Up">▲</span>
         <span class="dragger" @click="movePlaceDown(index)" title="Move Down">▼</span>
         <input v-model="place.label" placeholder="Label" style="max-width: 100px;" />
         <input v-model="place.lat" type="number" step="0.000001" placeholder="Latitude" />
         <input v-model="place.lon" type="number" step="0.000001" placeholder="Longitude" />
         <button class="danger" @click="removePlace(index)">X</button>
      </div>
    </div>

    <div class="section">
      <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom: 10px;">
          <h2 style="margin:0;">Solar Sites</h2>
          <button @click="addSite">Add Site</button>
      </div>

      <div v-if="!config.sites || config.sites.length === 0" style="color:#666; font-size: 13px; margin-bottom: 10px;">
        No solar sites yet.
      </div>

      <div v-for="(site, index) in config.sites" :key="site.id || index" class="site-item">
         <div class="site-field name-field">
           <label>Site Name</label>
           <input v-model="site.name" />
         </div>
         <div class="site-field id-field">
           <label>Site ID</label>
           <input v-model="site.id" />
         </div>
         <div class="site-field">
           <label>System kW</label>
           <input v-model.number="site.system_kw" type="number" step="0.01" min="0" />
         </div>
         <div class="site-field">
           <label>Efficiency</label>
           <input v-model.number="site.efficiency_factor" type="number" step="0.01" min="0" max="1" />
         </div>
         <div class="site-field">
           <label>Battery Capacity Ah</label>
           <input v-model.number="site.total_battery_capacity_ah" type="number" step="0.1" min="0" />
         </div>
         <div class="site-field">
           <label>Max Charging A</label>
           <input v-model.number="site.max_charging_ampere" type="number" step="1" min="0" />
         </div>
         <div class="site-field">
           <label>Initial SOC %</label>
           <input v-model.number="site.initial_soc_percent" type="number" step="1" min="0" max="100" />
         </div>
         <div class="site-field">
           <label>Stop Discharge SOC %</label>
           <input v-model.number="site.stop_discharge_soc_percent" type="number" step="1" min="0" max="100" />
         </div>
         <div class="site-field">
           <label>Consumption kWh / Hour</label>
           <input v-model.number="site.consumption_kwh_per_hour" type="number" step="0.1" min="0" />
         </div>
         <label class="site-check">
           <input v-model="site.charge_first" type="checkbox" />
           Charge first
         </label>
         <div class="site-field">
           <label>Latitude</label>
           <input v-model.number="site.lat" type="number" step="0.000001" />
         </div>
         <div class="site-field">
           <label>Longitude</label>
           <input v-model.number="site.lon" type="number" step="0.000001" />
         </div>
         <a :href="'/solar-site?site=' + encodeURIComponent(site.id || '')" target="_blank" style="font-size: 13px; font-weight: bold; color:#111;">Forecast</a>
         <button class="danger" @click="removeSite(index)">X</button>
      </div>
    </div>

    <div style="text-align: right; margin-top: 20px;">
       <button @click="saveConfig" style="font-size: 16px; padding: 10px 20px;">Save Configuration</button>
    </div>
  </div>
</div>

<script>
const { createApp } = Vue;

createApp({
  data() {
    return {
      loading: true,
      config: { default_list: "", lists: [], sites: [] },
      activeListId: null
    };
  },
  computed: {
    activeList() {
      if (!this.config.lists) return null;
      return this.config.lists.find(l => l.id === this.activeListId);
    }
  },
  mounted() {
    this.fetchConfig();
  },
  methods: {
    fetchConfig() {
      fetch('/api/places')
        .then(res => res.json())
        .then(data => {
          this.config = data;
          if (!this.config.lists) this.config.lists = [];
          if (!this.config.sites) this.config.sites = [];
          this.config.sites.forEach(site => {
            if (site.total_battery_capacity_ah === undefined) {
              site.total_battery_capacity_ah = site.total_battery_capacity_kwh === undefined ? 100 : site.total_battery_capacity_kwh;
            }
            if (site.max_charging_ampere === undefined) site.max_charging_ampere = 100;
            if (site.initial_soc_percent === undefined) site.initial_soc_percent = 20;
            if (site.stop_discharge_soc_percent === undefined) site.stop_discharge_soc_percent = 20;
            if (site.consumption_kwh_per_hour === undefined) site.consumption_kwh_per_hour = 0;
            if (site.charge_first === undefined) site.charge_first = false;
          });
          if (this.config.lists && this.config.lists.length > 0) {
              if(!this.activeListId || !this.config.lists.find(l=>l.id === this.activeListId)){
                  this.activeListId = this.config.lists[0].id;
              }
          }
          this.loading = false;
        });
    },
    saveConfig() {
      fetch('/api/places', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(this.config)
      })
      .then(res => {
         if (res.ok) {
             alert('Saved successfully!');
             this.fetchConfig();
         } else {
             alert('Error saving configuration.');
         }
      });
    },
    addList() {
      const newId = 'list_' + Date.now();
      this.config.lists.push({
        id: newId,
        name: 'New List',
        places: []
      });
      this.activeListId = newId;
    },
    removeList(index) {
      if (confirm('Are you sure you want to delete this list?')) {
        this.config.lists.splice(index, 1);
        if (this.config.lists.length > 0) {
            this.activeListId = this.config.lists[0].id;
        } else {
            this.activeListId = null;
        }
      }
    },
    addPlace() {
      if (this.activeList) {
        this.activeList.places.push({ label: '', lat: 0, lon: 0 });
      }
    },
    addSite() {
      if (!this.config.sites) this.config.sites = [];
      const newId = 'site_' + Date.now();
      this.config.sites.push({
        id: newId,
        name: 'New Solar Site',
        system_kw: 95.84,
        efficiency_factor: 0.80,
        total_battery_capacity_ah: 100,
        max_charging_ampere: 100,
        initial_soc_percent: 20,
        stop_discharge_soc_percent: 20,
        consumption_kwh_per_hour: 0,
        charge_first: false,
        lat: 13.174,
        lon: 121.278
      });
    },
    removeSite(index) {
      if (confirm('Are you sure you want to delete this site?')) {
        this.config.sites.splice(index, 1);
      }
    },
    removePlace(index) {
      if (this.activeList) {
        this.activeList.places.splice(index, 1);
      }
    },
    movePlaceUp(index) {
        if (index > 0 && this.activeList) {
            const temp = this.activeList.places[index - 1];
            this.activeList.places[index - 1] = this.activeList.places[index];
            this.activeList.places[index] = temp;
        }
    },
    movePlaceDown(index) {
        if (this.activeList && index < this.activeList.places.length - 1) {
            const temp = this.activeList.places[index + 1];
            this.activeList.places[index + 1] = this.activeList.places[index];
            this.activeList.places[index] = temp;
        }
    }
  }
}).mount('#app');
</script>
</body>
</html>"""
    return Response(html, mimetype="text/html")


@app.get("/api/places")
def api_get_places():
    if not os.path.exists(DEFAULT_PLACES_FILE):
        return {"default_list": "", "lists": [], "sites": []}
    return read_config_file(DEFAULT_PLACES_FILE)


@app.post("/api/places")
def api_post_places():
    data = request.json
    if not data:
        return Response("Invalid JSON", status=400)

    if "lists" not in data:
        data["lists"] = []
    if "sites" not in data:
        data["sites"] = []

    # Save to file
    with open(DEFAULT_PLACES_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

    return {"status": "ok"}


if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
